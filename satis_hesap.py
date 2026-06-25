import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
import datetime
import os
import sqlite3
import shutil
import json
import socket
import threading

# requests isteğe bağlı (İstemci modu için)
try:
    import requests as _requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

# â”€â”€â”€ Yapılandırma â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
_CONFIG_FILE = os.path.join(_BASE_DIR, "config.json")

def load_config() -> dict:
    defaults = {
        "mod": "yerel",          # yerel | istemci
        "db_yolu": os.path.join(_BASE_DIR, "satis_takip.db"),
        "sunucu_url": "http://127.0.0.1:8765"
    }
    if os.path.exists(_CONFIG_FILE):
        try:
            with open(_CONFIG_FILE, "r", encoding="utf-8") as f:
                stored = json.load(f)
            defaults.update(stored)
        except Exception:
            pass
    return defaults

def save_config(cfg: dict):
    with open(_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def db_connect():
    """Yerel modda SQLite bağlantısı döner."""
    cfg = load_config()
    path = cfg.get("db_yolu", os.path.join(_BASE_DIR, "satis_takip.db"))
    conn = sqlite3.connect(path, timeout=15)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def api_call(method: str, endpoint: str, **kwargs):
    """İstemci modunda HTTP API çağrısı yapar."""
    if not REQUESTS_OK:
        raise RuntimeError("'requests' kütüphanesi kurulu değil.\nKurmak için: pip install requests")
    cfg = load_config()
    url = cfg.get("sunucu_url", "http://127.0.0.1:8765").rstrip("/") + endpoint
    resp = getattr(_requests, method.lower())(url, timeout=10, **kwargs)
    resp.raise_for_status()
    return resp.json()

def is_client_mode() -> bool:
    return load_config().get("mod") == "istemci"

def is_share_mode() -> bool:
    return load_config().get("mod") == "paylasim"


# --- ReportLab Türkçe Karakter Font Kaydı ---
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

FONT_NAME = "Helvetica"
FONT_NAME_BOLD = "Helvetica-Bold"

try:
    arial_path = "C:/Windows/Fonts/arial.ttf"
    arial_bold_path = "C:/Windows/Fonts/arialbd.ttf"
    if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
        pdfmetrics.registerFont(TTFont('Arial', arial_path))
        pdfmetrics.registerFont(TTFont('Arial-Bold', arial_bold_path))
        FONT_NAME = "Arial"
        FONT_NAME_BOLD = "Arial-Bold"
except Exception:
    pass

# --- Uygulama İkonu Oluşturma (SVG Logo â†’ .ico) ---
def create_app_icon():
    """Web sitesi SVG logosunu Pillow ile .ico dosyasına çevirir."""
    try:
        from PIL import Image, ImageDraw
        sizes = [16, 32, 48, 64, 128, 256]
        images = []
        for size in sizes:
            img = Image.new('RGBA', (size, size), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            scale = size / 32.0

            # Koyu arka plan â€” #1E1E2E, yuvarlatılmış köşeler
            radius = max(1, int(8 * scale))
            draw.rounded_rectangle([0, 0, size - 1, size - 1], radius=radius, fill=(30, 30, 46, 255))

            lw = max(1, int(3 * scale))
            purple = (108, 92, 231, 255)   # #6C5CE7
            cyan   = (0, 210, 255, 255)    # #00D2FF

            def draw_chevron(pts, color):
                # Çizgileri çiz
                for i in range(len(pts) - 1):
                    x1, y1 = pts[i]
                    x2, y2 = pts[i + 1]
                    draw.line([(x1, y1), (x2, y2)], fill=color, width=lw)
                # Yuvarlak uçlar ve köşe doldurma
                r = max(0, lw // 2)
                for (x, y) in pts:
                    draw.ellipse([x - r, y - r, x + r, y + r], fill=color)

            # Mor şerit: M8 8 L16 16 L8 24
            draw_chevron(
                [(8 * scale, 8 * scale), (16 * scale, 16 * scale), (8 * scale, 24 * scale)],
                purple
            )
            # Cyan şerit: M16 8 L24 16 L16 24
            draw_chevron(
                [(16 * scale, 8 * scale), (24 * scale, 16 * scale), (16 * scale, 24 * scale)],
                cyan
            )
            images.append(img)

        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hausmart_icon.ico")
        images[0].save(
            icon_path, format='ICO',
            sizes=[(s, s) for s in sizes],
            append_images=images[1:]
        )
        return icon_path
    except Exception:
        return None

# --- Veritabanı Kurulumu ---
def init_db():
    if is_client_mode():
        return  # Sunucu kendi DB'sini yonetir
    conn = db_connect()
    cursor = conn.cursor()

    # Kullanıcılar tablosu
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS kullanicilar (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kullanici_adi TEXT UNIQUE NOT NULL,
        sifre TEXT NOT NULL,
        rol TEXT NOT NULL,
        aktif INTEGER DEFAULT 1
    )
    """)

    # Eski veritabanlarına 'aktif' kolonu ekle (migration)
    try:
        cursor.execute("ALTER TABLE kullanicilar ADD COLUMN aktif INTEGER DEFAULT 1")
        conn.commit()
    except Exception:
        pass

    # Satışlar tablosu
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS satislar (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT NOT NULL,
        kullanici TEXT NOT NULL,
        musteri_adi TEXT,
        urun_adi TEXT,
        miktar REAL,
        birim TEXT,
        fiyat_birimi TEXT,
        torba_agirligi REAL,
        alis_fiyati REAL,
        baz_satis_fiyati REAL,
        odeme_turu TEXT,
        vade_ay INTEGER,
        vade_orani REAL,
        birim_fiyat REAL,
        toplam_tutar REAL,
        kar REAL,
        irsaliye_yolu TEXT
    )
    """)

    # Varsayılan kullanıcıları ekle
    defaults = [
        ("superadmin", "super123",  "Süper Admin"),
        ("admin",      "admin123",  "Yönetici"),
        ("satis",      "satis123",  "Personel"),
    ]
    for adi, sifre, rol in defaults:
        try:
            cursor.execute(
                "INSERT INTO kullanicilar (kullanici_adi, sifre, rol, aktif) VALUES (?, ?, ?, 1)",
                (adi, sifre, rol)
            )
        except sqlite3.IntegrityError:
            pass
    conn.commit()
    conn.close()

# --- Satış Detay Penceresi ---
class SaleDetailsDialog(tk.Toplevel):
    def __init__(self, parent, sale_id, role):
        super().__init__(parent)
        self.title(f"Satış Detayı - ID: {sale_id}")
        self.geometry("460x520")
        self.resizable(False, False)
        self.configure(bg="#f5f5f7")
        self.transient(parent)
        self.grab_set()
        
        self.sale_id = sale_id
        self.role = role
        
        self.center_window(460, 520)
        self.create_widgets()
        
    def center_window(self, width, height):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        self.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
        
    def create_widgets(self):
        if is_client_mode():
            try:
                data = api_call("get", f"/api/satislar/{self.sale_id}")
                row = (
                    data["tarih"], data["kullanici"], data.get("musteri_adi"), data.get("urun_adi"),
                    data.get("miktar"), data.get("birim"), data.get("fiyat_birimi"), data.get("torba_agirligi"),
                    data.get("alis_fiyati", 0.0), data.get("baz_satis_fiyati", 0.0), data.get("odeme_turu"),
                    data.get("vade_ay", 0), data.get("vade_orani", 0.0), data.get("birim_fiyat", 0.0),
                    data.get("toplam_tutar", 0.0), data.get("kar", 0.0), data.get("irsaliye_yolu", "")
                )
            except Exception as e:
                messagebox.showerror("Hata", f"Satış bilgisi alınamadı: {e}")
                self.destroy()
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("""
            SELECT tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi,
                   alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar, irsaliye_yolu
            FROM satislar WHERE id = ?
            """, (self.sale_id,))
            row = cursor.fetchone()
            conn.close()
        
        if not row:
            ttk.Label(self, text="Kayıt bulunamadı.").pack(padx=20, pady=20)
            return
            
        tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi, \
        alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar, irsaliye_yolu = row
        
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(main_frame, text=f"Satış Detayı (ID: {self.sale_id})", font=("Helvetica", 12, "bold"), foreground="#1d1d1f").pack(pady=(0, 10))
        
        grid_frame = ttk.LabelFrame(main_frame, text="Detaylar")
        grid_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        details = [
            ("Tarih:", tarih),
            ("Satışı Yapan:", kullanici),
            ("Müşteri:", musteri_adi),
            ("Ürün Adı:", urun_adi),
            ("Miktar:", f"{miktar:g} {birim}"),
            ("Satış Fiyat Birimi:", fiyat_birimi),
            ("Ödeme Türü:", odeme_turu),
            ("Vade:", f"{vade_ay} Ay (Aylık %{vade_orani:.2f})" if vade_ay > 0 else "Nakit / Vadesiz"),
            ("Birim Fiyat:", f"{birim_fiyat:,.2f} â‚º"),
            ("Toplam Tutar:", f"{toplam_tutar:,.2f} â‚º")
        ]
        
        if (birim == "TORBA" or fiyat_birimi == "TORBA"):
            details.insert(6, ("Torba Ağırlığı:", f"{torba_agirligi:g} kg"))
            
        if self.role == "Yönetici":
            details.append(("Alış Fiyatı:", f"{alis_fiyati:,.2f} â‚º"))
            details.append(("Toplam KÃ¢r:", f"{kar:,.2f} â‚º"))
            
        details.append(("İrsaliye Durumu:", "Yüklendi" if irsaliye_yolu else "Yüklenmedi"))
        
        for i, (label_txt, val_txt) in enumerate(details):
            ttk.Label(grid_frame, text=label_txt, font=("Helvetica", 9, "bold"), foreground="#1d1d1f").grid(row=i, column=0, sticky=tk.W, padx=5, pady=3)
            ttk.Label(grid_frame, text=val_txt, font=("Helvetica", 9)).grid(row=i, column=1, sticky=tk.W, padx=15, pady=3)
            
        ttk.Button(main_frame, text="Kapat", command=self.destroy).pack(side=tk.RIGHT)

# --- Ana Uygulama ---
class HausmartApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Hausmart Fiyatlandırma & Satış Takip Sistemi")
        self.root.configure(bg="#f5f5f7")

        # Uygulama ikonunu ayarla (web sitesi SVG logo)
        icon_path = create_app_icon()
        if icon_path and os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass

        # Veritabanını kur
        init_db()

        # Giriş ekranını göster
        self.show_login_frame()
        
    def center_window(self, width, height):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        self.root.geometry(f'{width}x{height}+{int(x)}+{int(y)}')

    def show_login_frame(self):
        for widget in self.root.winfo_children():
            widget.destroy()
            
        self.root.geometry("400x560")
        self.center_window(400, 560)
        self.root.resizable(False, False)
        
        self.login_frame = ttk.Frame(self.root)
        self.login_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        ttk.Label(self.login_frame, text="Hausmart Giriş Paneli", font=("Helvetica", 14, "bold"), foreground="#1d1d1f").pack(pady=(0, 20))
        
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        
        cfg = load_config()
        self.server_ip_var = tk.StringVar(value=cfg.get("sunucu_url", ""))
        self.db_path_var = tk.StringVar(value=cfg.get("db_yolu", ""))
        
        ttk.Label(self.login_frame, text="Kullanıcı Adı:").pack(anchor=tk.W, pady=(0,5))
        self.username_entry = ttk.Entry(self.login_frame, textvariable=self.username_var, width=30)
        self.username_entry.pack(pady=(0,10))
        
        ttk.Label(self.login_frame, text="Şifre:").pack(anchor=tk.W, pady=(0,5))
        entry_pass = ttk.Entry(self.login_frame, textvariable=self.password_var, show="*", width=30)
        entry_pass.pack(pady=(0,10))
        entry_pass.bind("<Return>", lambda e: self.perform_login())
        
        ttk.Label(self.login_frame, text="Sunucu Adresi (İstemci ise):").pack(anchor=tk.W, pady=(0,5))
        ttk.Entry(self.login_frame, textvariable=self.server_ip_var, width=30).pack(pady=(0,10))
        
        ttk.Label(self.login_frame, text="Veritabanı Yolu (Yerel/Ağ ise):").pack(anchor=tk.W, pady=(0,5))
        ttk.Entry(self.login_frame, textvariable=self.db_path_var, width=30).pack(pady=(0,15))
        
        btn_frame = ttk.Frame(self.login_frame)
        btn_frame.pack(fill=tk.X, pady=5)
 
        ttk.Button(btn_frame, text="Giriş Yap", command=self.perform_login).pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text="⚙ Bağlantı", command=self.show_connection_dialog).pack(side=tk.LEFT)
 
        mod_text = ("🌐 İstemci: " + cfg.get("sunucu_url","")) if cfg.get("mod") == "istemci" else "🖥 Yerel Mod"
        self.lbl_conn_mode = ttk.Label(self.login_frame, text=mod_text,
                                       font=("Helvetica", 8), foreground="#0071e3")
        self.lbl_conn_mode.pack(pady=(8,0))

        # Update widgets
        self.update_frame = ttk.Frame(self.login_frame)
        self.lbl_update_status = ttk.Label(self.update_frame, text="", font=("Helvetica", 9), foreground="#0071e3")
        self.lbl_update_status.pack(pady=(5, 2))
        self.btn_update = ttk.Button(self.update_frame, text="Son Sürüme Güncelle", command=self.install_python_update)
        
        # Version label
        ttk.Label(self.login_frame, text="Sürüm: v1.0.0", font=("Helvetica", 8), foreground="#aaaaaa").pack(pady=(5,0))
        
        # Start background check thread
        threading.Thread(target=self.check_python_update, daemon=True).start()

    def perform_login(self):
        cfg = load_config()
        server_url = self.server_ip_var.get().strip()
        db_path = self.db_path_var.get().strip()
        
        changed = False
        if server_url:
            if not server_url.startswith("http://") and not server_url.startswith("https://"):
                server_url = "http://" + server_url
            url_part = server_url[7:] if server_url.startswith("http://") else server_url[8:]
            if ":" not in url_part:
                server_url = server_url.rstrip("/") + ":8765"
            
            if cfg.get("mod") != "istemci" or cfg.get("sunucu_url") != server_url:
                cfg["mod"] = "istemci"
                cfg["sunucu_url"] = server_url
                changed = True
        elif db_path:
            target_mod = "paylasim" if (db_path.startswith("\\\\") or db_path.startswith("//")) else "yerel"
            if cfg.get("mod") != target_mod or cfg.get("db_yolu") != db_path:
                cfg["mod"] = target_mod
                cfg["db_yolu"] = db_path
                changed = True
        
        if changed:
            save_config(cfg)
            mod_text = ("🌐 İstemci: " + cfg.get("sunucu_url","")) if cfg.get("mod") == "istemci" else "🖥 Yerel Mod"
            self.lbl_conn_mode.config(text=mod_text)
            
        u = self.username_var.get()
        p = self.password_var.get()

        if is_client_mode():
            try:
                data = api_call("post", "/api/login",
                                json={"kullanici_adi": u, "sifre": p})
                row = (data["kullanici_adi"], data["rol"], data.get("aktif", 1))
            except Exception as _e:
                _msg = str(_e)
                if "401" in _msg or "403" in _msg:
                    messagebox.showerror("Hata",
                        "Sunucu: Gecersiz kimlik bilgileri veya hesap pasif!")
                else:
                    messagebox.showerror("Baglanti Hatasi",
                        f"Sunucuya baglanamadı:\n{_e}\n\n"
                        "Baglantiyi test etmek icin giris ekranindaki "
                        "Baglanti butonunu kullanin.")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT kullanici_adi, rol, aktif FROM kullanicilar WHERE kullanici_adi = ? AND sifre = ?",
                (u, p)
            )
            row = cursor.fetchone()
            conn.close()

        if row:
            if row[2] == 0:
                messagebox.showerror("Erişim Engellendi",
                    "Bu hesap pasif durumda.\nLütfen yöneticinizle iletişime geçin.")
                return
            self.current_user = row[0]
            self.current_role = row[1]
            
            self.login_frame.destroy()
            
            self.root.resizable(True, True)
            self.root.geometry("1150x850")
            self.center_window(1150, 850)
            
            style = ttk.Style()
            style.theme_use('clam')
            style.configure("TFrame", background="#f5f5f7")
            style.configure("TLabelframe", background="#ffffff", padding=10)
            style.configure("TLabelframe.Label", background="#ffffff", font=("Helvetica", 11, "bold"), foreground="#1d1d1f")
            style.configure("TLabel", background="#ffffff", font=("Helvetica", 10), foreground="#86868b")
            style.configure("TButton", font=("Helvetica", 10, "bold"), padding=6)
            style.configure("Treeview", font=("Helvetica", 10), rowheight=25)
            style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))
            
            # --- Değişkenler ---
            self.cust_name = tk.StringVar()
            self.prod_name = tk.StringVar()
            self.unit_type = tk.StringVar(value="TORBA")
            self.qty = tk.StringVar(value="1")
            self.purchase_price = tk.StringVar(value="")
            self.base_price = tk.StringVar(value="")
            self.receipt_type = tk.StringVar(value="Nakit")
            self.purchase_price_type = tk.StringVar(value="TORBA")
            self.base_price_type = tk.StringVar(value="TON")
            self.bag_weight = tk.StringVar(value="50")
            
            self.vade_months = tk.StringVar(value="0")
            self.vade_rate = tk.StringVar(value="0")
            
            self.has_shipping = tk.BooleanVar(value=False)
            self.shipping_cost = tk.StringVar(value="")
            self.has_unloading = tk.BooleanVar(value=False)
            self.unloading_cost = tk.StringVar(value="")
            
            self.rate_cash = tk.StringVar(value="0")
            self.rate_cc = tk.StringVar(value="3")
            self.rate_check = tk.StringVar(value="5")
            self.rate_note = tk.StringVar(value="8")
            self.rate_doc = tk.StringVar(value="4")
            self.rate_dbs = tk.StringVar(value="2")

            self.calculated_data = {}

            vars_to_trace = [
                self.qty, self.purchase_price, self.base_price, self.vade_months, self.vade_rate,
                self.has_shipping, self.shipping_cost, self.has_unloading, self.unloading_cost,
                self.rate_cash, self.rate_cc, self.rate_check, self.rate_note, self.rate_doc, self.rate_dbs,
                self.receipt_type, self.cust_name, self.prod_name, self.unit_type,
                self.purchase_price_type, self.base_price_type, self.bag_weight
            ]
            for v in vars_to_trace:
                v.trace_add("write", lambda *args: self.calculate())
                
            self.create_notebook()
        else:
            messagebox.showerror("Hata", "Geçersiz kullanıcı adı veya şifre!")

    def create_notebook(self):
        header_frame = ttk.Frame(self.root)
        header_frame.pack(fill=tk.X, padx=15, pady=5)
        
        user_info = f"Giriş Yapan: {self.current_user} ({self.current_role})"
        ttk.Label(header_frame, text=user_info, font=("Helvetica", 10, "bold"), foreground="#0071e3").pack(side=tk.RIGHT)
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        self.tab1 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)
        self.tab3 = ttk.Frame(self.notebook)

        self.notebook.add(self.tab1, text="Fiyat Hesaplama ve Satış")
        self.notebook.add(self.tab2, text="Son Hareketler & İrsaliye")
        self.notebook.add(self.tab3, text="Ürün Fiyat Geçmişi")

        self.tab4 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab4, text="⚙ Kullanıcı Yönetimi")

        if self.current_role == "Süper Admin":
            self.tab5 = ttk.Frame(self.notebook)
            self.notebook.add(self.tab5, text="Sunucu Yonetimi")

        self.create_tab1_widgets()
        self.create_tab2_widgets()
        self.create_tab3_widgets()
        self.create_tab4_widgets()
        if self.current_role == "Süper Admin":
            self.create_tab5_widgets()

        self.load_payment_rates()
        self.periodic_rate_sync()

        self.calculate()

    # --- Sekme 1: Hesaplama & Satış Ekranı ---
    def create_tab1_widgets(self):
        main_frame = ttk.Frame(self.tab1)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        lf1 = ttk.LabelFrame(main_frame, text="Ürün ve Müşteri Bilgileri")
        lf1.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(lf1, text="Müşteri Adı / Unvanı:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(lf1, textvariable=self.cust_name, width=25).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(lf1, text="Ürün Adı / Kodu:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(lf1, textvariable=self.prod_name, width=25).grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(lf1, text="Miktar Birimi:").grid(row=0, column=4, sticky=tk.W, padx=5, pady=5)
        ttk.Combobox(lf1, textvariable=self.unit_type, values=["TORBA", "KG", "TON", "M2"], width=10, state="readonly").grid(row=0, column=5, padx=5, pady=5)

        ttk.Label(lf1, text="Miktar:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(lf1, textvariable=self.qty, width=25).grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(lf1, text="Alış Fiyatı (Opsiyonel):").grid(row=1, column=2, sticky=tk.W, padx=5, pady=5)
        self.ent_purchase_price = ttk.Entry(lf1, textvariable=self.purchase_price, width=25)
        self.ent_purchase_price.grid(row=1, column=3, padx=5, pady=5)
        if self.current_role != "Yönetici":
            self.ent_purchase_price.config(state="disabled")

        ttk.Label(lf1, text="Baz Satış Fiyatı (*):").grid(row=1, column=4, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(lf1, textvariable=self.base_price, width=15).grid(row=1, column=5, padx=5, pady=5)

        ttk.Label(lf1, text="Fişe Yansıyacak Tür:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        receipt_cb = ttk.Combobox(lf1, textvariable=self.receipt_type, values=["Nakit", "Kredi Kartı", "Çek", "Senet", "Evrak", "DBS"], state="readonly", width=23)
        receipt_cb.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(lf1, text="Alış Fiyat Birimi:").grid(row=2, column=2, sticky=tk.W, padx=5, pady=5)
        self.cb_purchase_price_type = ttk.Combobox(lf1, textvariable=self.purchase_price_type, values=["TORBA", "KG", "TON", "M2"], width=23, state="readonly")
        self.cb_purchase_price_type.grid(row=2, column=3, padx=5, pady=5)
        if self.current_role != "Yönetici":
            self.cb_purchase_price_type.config(state="disabled")

        ttk.Label(lf1, text="Satış Fiyat Birimi:").grid(row=2, column=4, sticky=tk.W, padx=5, pady=5)
        ttk.Combobox(lf1, textvariable=self.base_price_type, values=["TORBA", "KG", "TON", "M2"], width=15, state="readonly").grid(row=2, column=5, padx=5, pady=5)

        ttk.Label(lf1, text="Torba Ağırlığı (kg):").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.ent_bag_weight = ttk.Entry(lf1, textvariable=self.bag_weight, width=25)
        self.ent_bag_weight.grid(row=3, column=1, padx=5, pady=5)

        self.lbl_conversion_info = ttk.Label(lf1, text="", font=("Helvetica", 9, "italic"), foreground="#0071e3")
        self.lbl_conversion_info.grid(row=3, column=2, columnspan=4, sticky=tk.W, padx=5, pady=5)

        # 2. Vade & Ek Maliyetler
        middle_frame = ttk.Frame(main_frame)
        middle_frame.pack(fill=tk.X, pady=(0, 15))

        lf2 = ttk.LabelFrame(middle_frame, text="Vadeli Satış Seçenekleri (Nakit Hariç)")
        lf2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        ttk.Label(lf2, text="Vade Süresi (Ay):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Combobox(lf2, textvariable=self.vade_months, values=[str(i) for i in range(13)], state="readonly", width=10).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(lf2, text="Aylık Vade Farkı (%):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(lf2, textvariable=self.vade_rate, width=13).grid(row=1, column=1, padx=5, pady=5)

        self.lbl_vade_summary = ttk.Label(lf2, text="Ek Vade Farkı Yükü: %0.00", font=("Helvetica", 10, "bold"), foreground="#0071e3")
        self.lbl_vade_summary.grid(row=2, column=0, columnspan=2, pady=10)

        lf3 = ttk.LabelFrame(middle_frame, text="Ek Hizmetler (Toplam Tutara Eklenir)")
        lf3.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        ttk.Checkbutton(lf3, text="Nakliye Dahil", variable=self.has_shipping).grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.ent_ship = ttk.Entry(lf3, textvariable=self.shipping_cost, width=15)
        self.ent_ship.grid(row=0, column=1, padx=5, pady=5)

        ttk.Checkbutton(lf3, text="İndirme Dahil", variable=self.has_unloading).grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.ent_unload = ttk.Entry(lf3, textvariable=self.unloading_cost, width=15)
        self.ent_unload.grid(row=1, column=1, padx=5, pady=5)

        # 3. Ödeme Türü Baz Oranları
        lf4 = ttk.LabelFrame(main_frame, text="Ödeme Türü Baz Oranları (%)")
        lf4.pack(fill=tk.X, pady=(0, 15))
        
        rates_info = [("Nakit:", self.rate_cash), ("Kredi Kartı:", self.rate_cc), ("Çek:", self.rate_check),
                      ("Senet:", self.rate_note), ("Evrak:", self.rate_doc), ("DBS:", self.rate_dbs)]
        
        self.rate_widgets = {}
        for i, (lbl, var) in enumerate(rates_info):
            key_name = ["cash", "cc", "check", "note", "doc", "dbs"][i]
            ttk.Label(lf4, text=lbl).grid(row=0, column=i*2, sticky=tk.W, padx=5, pady=10)
            entry = ttk.Entry(lf4, textvariable=var, width=8)
            entry.grid(row=0, column=i*2+1, padx=(0, 15), pady=10)
            entry.bind("<FocusOut>", lambda e: self.save_rates())
            entry.bind("<Return>", lambda e: self.save_rates())
            self.rate_widgets[key_name] = entry

        # 4. Sonuç Tablosu
        lf5 = ttk.LabelFrame(main_frame, text="Kümülatif Fiyat Analizi")
        lf5.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        self.columns = ("tur", "baz_oran", "kum_oran", "birim_fiyat", "toplam_tutar", "kar")

        self.tree = ttk.Treeview(lf5, columns=self.columns, show="headings", height=6)
        self.tree.heading("tur", text="Ödeme Türü")
        self.tree.heading("baz_oran", text="Baz Oran (%)")
        self.tree.heading("kum_oran", text="Kümülatif Oran (%)")
        self.tree.heading("birim_fiyat", text="1 Birim Fiyatı (₺)")
        self.tree.heading("toplam_tutar", text="Toplam Tutar (₺)")
        self.tree.heading("kar", text="Tahmini Toplam Kâr (₺)")
        
        self.tree.column("tur", anchor=tk.W, width=150)
        self.tree.column("baz_oran", anchor=tk.CENTER, width=100)
        self.tree.column("kum_oran", anchor=tk.CENTER, width=150)
        self.tree.column("birim_fiyat", anchor=tk.E, width=150)
        self.tree.column("toplam_tutar", anchor=tk.E, width=150)

        self.tree.column("kar", anchor=tk.E, width=150)

        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 5. Butonlar
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(btn_frame, text="Excel'e Aktar", command=self.export_excel).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Satışı Kaydet", command=self.save_sale_only).pack(side=tk.RIGHT, padx=10)
        ttk.Button(btn_frame, text="Satış Çıktısı Al (PDF)", command=self.export_pdf_only).pack(side=tk.RIGHT, padx=10)

    # --- Sekme 2: Son Hareketler & İrsaliye Ekranı ---
    def create_tab2_widgets(self):
        main_frame = ttk.Frame(self.tab2)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(top_frame, text="Son Satış Hareketleri", font=("Helvetica", 12, "bold"), foreground="#1d1d1f").pack(side=tk.LEFT)
        ttk.Button(top_frame, text="Yenile", command=self.load_recent_movements).pack(side=tk.RIGHT)
        
        lf_list = ttk.LabelFrame(main_frame, text="Kayıtlar")
        lf_list.pack(fill=tk.BOTH, expand=True)
        
        if self.current_role == "Yönetici":
            self.mov_columns = ("id", "tarih", "kullanici", "musteri", "urun", "miktar", "birim", "tutar", "kar", "irsaliye")
        else:
            self.mov_columns = ("id", "tarih", "kullanici", "musteri", "urun", "miktar", "birim", "tutar", "irsaliye")
            
        self.mov_tree = ttk.Treeview(lf_list, columns=self.mov_columns, show="headings", height=15)
        
        self.mov_tree.heading("id", text="ID")
        self.mov_tree.heading("tarih", text="Tarih")
        self.mov_tree.heading("kullanici", text="Kullanıcı")
        self.mov_tree.heading("musteri", text="Müşteri")
        self.mov_tree.heading("urun", text="Ürün")
        self.mov_tree.heading("miktar", text="Miktar")
        self.mov_tree.heading("birim", text="Birim")
        self.mov_tree.heading("tutar", text="Toplam Tutar (â‚º)")
        self.mov_tree.heading("irsaliye", text="İrsaliye Durumu")
        
        self.mov_tree.column("id", anchor=tk.CENTER, width=50)
        self.mov_tree.column("tarih", anchor=tk.CENTER, width=130)
        self.mov_tree.column("kullanici", anchor=tk.CENTER, width=100)
        self.mov_tree.column("musteri", anchor=tk.W, width=150)
        self.mov_tree.column("urun", anchor=tk.W, width=150)
        self.mov_tree.column("miktar", anchor=tk.CENTER, width=80)
        self.mov_tree.column("birim", anchor=tk.CENTER, width=80)
        self.mov_tree.column("tutar", anchor=tk.E, width=120)
        self.mov_tree.column("irsaliye", anchor=tk.CENTER, width=120)
        
        if self.current_role == "Yönetici":
            self.mov_tree.heading("kar", text="KÃ¢r (â‚º)")
            self.mov_tree.column("kar", anchor=tk.E, width=100)
            
        self.mov_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="Satışı Görüntüle", command=self.show_sale_details).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="İrsaliye Yükle", command=self.upload_waybill).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="İrsaliyeyi Görüntüle", command=self.view_waybill).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="PDF Fişi Tekrar Çıkar", command=self.regenerate_pdf_from_history).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Detay PDF Çıkar", command=self.export_sale_detail_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Excel Çıktısı Al", command=self.export_sale_detail_excel).pack(side=tk.LEFT, padx=5)
        
        if self.current_role == "Yönetici":
            ttk.Button(btn_frame, text="Satışı Sil", command=self.delete_sale).pack(side=tk.RIGHT, padx=5)
            
        self.load_recent_movements()

    # --- Sekme 3: Ürün Fiyat Geçmişi Ekranı ---
    def create_tab3_widgets(self):
        main_frame = ttk.Frame(self.tab3)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(top_frame, text="Ürün Geçmiş Fiyat Trendi", font=("Helvetica", 12, "bold"), foreground="#1d1d1f").grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))
        
        ttk.Label(top_frame, text="Ürün Seçin:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5))
        self.cb_history_products = ttk.Combobox(top_frame, width=30, state="readonly")
        self.cb_history_products.grid(row=1, column=1, padx=(0, 10))
        self.cb_history_products.bind("<<ComboboxSelected>>", lambda e: self.load_price_history())
        
        ttk.Button(top_frame, text="Listeyi Yenile", command=self.update_history_products_combobox).grid(row=1, column=2)
        
        lf_hist = ttk.LabelFrame(main_frame, text="Geçmiş Satış Fiyatları")
        lf_hist.pack(fill=tk.BOTH, expand=True)
        
        self.hist_columns = ("tarih", "musteri", "miktar", "birim", "odeme_turu", "vade", "birim_fiyat", "toplam_tutar")
        self.hist_tree = ttk.Treeview(lf_hist, columns=self.hist_columns, show="headings", height=15)
        
        self.hist_tree.heading("tarih", text="Satış Tarihi")
        self.hist_tree.heading("musteri", text="Müşteri")
        self.hist_tree.heading("miktar", text="Miktar")
        self.hist_tree.heading("birim", text="Miktar Birimi")
        self.hist_tree.heading("odeme_turu", text="Ödeme Türü")
        self.hist_tree.heading("vade", text="Vade (Ay)")
        self.hist_tree.heading("birim_fiyat", text="Birim Satış Fiyatı (â‚º)")
        self.hist_tree.heading("toplam_tutar", text="Toplam Tutar (â‚º)")
        
        self.hist_tree.column("tarih", anchor=tk.CENTER, width=130)
        self.hist_tree.column("musteri", anchor=tk.W, width=180)
        self.hist_tree.column("miktar", anchor=tk.CENTER, width=80)
        self.hist_tree.column("birim", anchor=tk.CENTER, width=80)
        self.hist_tree.column("odeme_turu", anchor=tk.CENTER, width=100)
        self.hist_tree.column("vade", anchor=tk.CENTER, width=80)
        self.hist_tree.column("birim_fiyat", anchor=tk.E, width=150)
        self.hist_tree.column("toplam_tutar", anchor=tk.E, width=150)
        
        self.hist_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.update_history_products_combobox()

    # --- Yardımcı Fonksiyonlar ---
    def get_float(self, var):
        try:
            val = float(var.get().replace(',', '.'))
            return val if val >= 0 else 0.0
        except ValueError:
            return 0.0

    def calculate(self):
        base_price = self.get_float(self.base_price)
        purchase_price = self.get_float(self.purchase_price)
        qty = self.get_float(self.qty)
        if qty <= 0: qty = 1.0

        u_qty_type = self.unit_type.get()
        u_pur_type = self.purchase_price_type.get()
        u_sel_type = self.base_price_type.get()
        bag_w = self.get_float(self.bag_weight)
        if bag_w <= 0: bag_w = 50.0

        if "TORBA" in (u_qty_type, u_pur_type, u_sel_type):
            if hasattr(self, 'ent_bag_weight'):
                self.ent_bag_weight.state(['!disabled'])
        else:
            if hasattr(self, 'ent_bag_weight'):
                self.ent_bag_weight.state(['disabled'])

        weights = {
            "KG": 1.0,
            "TON": 1000.0,
            "TORBA": bag_w,
            "M2": 1.0
        }

        # Alış katsayısı (Alış Biriminden Miktar Birimine)
        pur_conversion_factor = 1.0
        if u_qty_type != u_pur_type:
            if u_qty_type != "M2" and u_pur_type != "M2":
                w_q = weights.get(u_qty_type, 1.0)
                w_pur = weights.get(u_pur_type, 1.0)
                pur_conversion_factor = w_q / w_pur

        # Satış katsayısı (Satış Biriminden Miktar Birimine)
        sel_conversion_factor = 1.0
        if u_qty_type != u_sel_type:
            if u_qty_type != "M2" and u_sel_type != "M2":
                w_q = weights.get(u_qty_type, 1.0)
                w_sel = weights.get(u_sel_type, 1.0)
                sel_conversion_factor = w_q / w_sel

        converted_purchase_price = purchase_price * pur_conversion_factor
        converted_base_price = base_price * sel_conversion_factor

        if hasattr(self, 'lbl_conversion_info'):
            info_text = []
            if u_qty_type != u_pur_type and u_qty_type != "M2" and u_pur_type != "M2" and purchase_price > 0:
                info_text.append(f"Maliyet Dönüşümü: 1 {u_qty_type} = {pur_conversion_factor:.4f} {u_pur_type} ({converted_purchase_price:,.2f} â‚º/{u_qty_type})")
            if u_qty_type != u_sel_type and u_qty_type != "M2" and u_sel_type != "M2" and base_price > 0:
                info_text.append(f"Satış Dönüşümü: 1 {u_qty_type} = {sel_conversion_factor:.4f} {u_sel_type} ({converted_base_price:,.2f} â‚º/{u_qty_type})")
            
            if info_text:
                self.lbl_conversion_info.config(text="  |  ".join(info_text))
            else:
                self.lbl_conversion_info.config(text="")

        vade_months = self.get_float(self.vade_months)
        vade_rate = self.get_float(self.vade_rate)
        total_vade_surcharge = vade_months * vade_rate

        if hasattr(self, 'lbl_vade_summary'):
            self.lbl_vade_summary.config(text=f"Ek Vade Farkı Yükü: %{total_vade_surcharge:.2f}")

        ship_cost = self.get_float(self.shipping_cost) if self.has_shipping.get() else 0.0
        if hasattr(self, 'ent_ship'):
            if not self.has_shipping.get(): self.ent_ship.state(['disabled'])
            else: self.ent_ship.state(['!disabled'])

        unload_cost = self.get_float(self.unloading_cost) if self.has_unloading.get() else 0.0
        if hasattr(self, 'ent_unload'):
            if not self.has_unloading.get(): self.ent_unload.state(['disabled'])
            else: self.ent_unload.state(['!disabled'])

        unit_extra_cost = (ship_cost + unload_cost) / qty
        unit_base_price = converted_base_price + unit_extra_cost
        unit_cost = (converted_purchase_price + unit_extra_cost) if converted_purchase_price > 0 else 0.0

        payment_types = [
            ("Nakit", self.get_float(self.rate_cash)),
            ("Kredi Kartı", self.get_float(self.rate_cc)),
            ("Çek", self.get_float(self.rate_check)),
            ("Senet", self.get_float(self.rate_note)),
            ("Evrak", self.get_float(self.rate_doc)),
            ("DBS", self.get_float(self.rate_dbs))
        ]

        if hasattr(self, 'tree'):
            for item in self.tree.get_children():
                self.tree.delete(item)

        self.calculated_data.clear()

        if base_price <= 0:
            return

        for p_name, base_rate in payment_types:
            final_applied_rate = base_rate if p_name == "Nakit" else (base_rate + total_vade_surcharge)
            
            unit_final_price = unit_base_price * (1 + (final_applied_rate / 100))
            total_final_price = unit_final_price * qty
            total_profit = ((unit_base_price - unit_cost) * qty) if unit_cost > 0 else 0.0

            self.calculated_data[p_name] = {
                "unit_price": unit_final_price,
                "total_price": total_final_price,
                "profit": total_profit,
                "cumulative_rate": final_applied_rate
            }

            profit_margin_pct = (total_profit / (unit_cost * qty) * 100) if unit_cost > 0 else 0.0
            kar_text = f"{total_profit:,.2f} ₺ ({profit_margin_pct:+.2f}%)" if purchase_price > 0 else "-"
            
            if hasattr(self, 'tree'):
                self.tree.insert("", tk.END, values=(
                    p_name,
                    f"%{base_rate:.2f}",
                    f"%{final_applied_rate:.2f}",
                    f"{unit_final_price:,.2f}",
                    f"{total_final_price:,.2f}",
                    kar_text
                ))

    def export_excel(self):
        if not self.calculated_data:
            messagebox.showwarning("Uyarı", "Hesaplanmış bir veri bulunmuyor. Önce fiyat giriniz.")
            return
            
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"Fiyat_Listesi_{self.prod_name.get()}.xlsx")
        if not filepath: return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fiyat Analizi"

            # Başlık bilgileri
            header_fill = PatternFill(start_color="F8CD24", end_color="F8CD24", fill_type="solid")
            bold_font = Font(bold=True, size=11)
            title_font = Font(bold=True, size=13)
            center_align = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A1:F1")
            ws["A1"] = "HAUSMART - FİYAT ANALİZ RAPORU"
            ws["A1"].font = title_font
            ws["A1"].fill = header_fill
            ws["A1"].alignment = center_align

            info_rows = [
                ("Müşteri:", self.cust_name.get() or "-"),
                ("Ürün:", self.prod_name.get() or "-"),
                ("Miktar:", f"{self.qty.get()} {self.unit_type.get()}"),
                ("Satış Fiyat Birimi:", self.base_price_type.get()),
                ("Baz Satış Fiyatı:", f"{self.get_float(self.base_price):,.2f} â‚º"),
                ("Tarih:", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")),
            ]
            if self.current_role == "Yönetici" and self.get_float(self.purchase_price) > 0:
                info_rows.append(("Alış Fiyatı:", f"{self.get_float(self.purchase_price):,.2f} â‚º ({self.purchase_price_type.get()})"))

            for i, (lbl, val) in enumerate(info_rows, start=2):
                ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)
                ws.cell(row=i, column=2, value=val)

            data_start_row = len(info_rows) + 3
            headers = ["Ödeme Türü", "Baz Oran (%)", "Kümülatif Oran (%)", "1 Birim Fiyatı (â‚º)", "Toplam Tutar (â‚º)"]
            if self.current_role == "Yönetici":
                headers.append("Tahmini KÃ¢r (â‚º)")

            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=data_start_row, column=col_idx, value=h)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = center_align

            rate_map = {
                "Nakit": self.get_float(self.rate_cash),
                "Kredi Kartı": self.get_float(self.rate_cc),
                "Çek": self.get_float(self.rate_check),
                "Senet": self.get_float(self.rate_note),
                "Evrak": self.get_float(self.rate_doc),
                "DBS": self.get_float(self.rate_dbs),
            }

            for row_offset, (p_name, vals) in enumerate(self.calculated_data.items(), start=1):
                r = data_start_row + row_offset
                base_r = rate_map.get(p_name, 0.0)
                row_data = [p_name, base_r, vals["cumulative_rate"], vals["unit_price"], vals["total_price"]]
                if self.current_role == "Yönetici":
                    kar = vals["profit"] if self.get_float(self.purchase_price) > 0 else "-"
                    row_data.append(kar)
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=r, column=col_idx, value=val)
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00'
                    if row_offset % 2 == 0:
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            for col in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)

            wb.save(filepath)
            messagebox.showinfo("Başarılı", f"Veriler Excel'e kaydedildi:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel oluşturulurken hata:\n{str(e)}")

    # --- Satışı Kaydet ---
    def save_sale_only(self):
        if not self.calculated_data:
            messagebox.showwarning("Uyarı", "Hesaplanmış bir veri bulunmuyor. Önce fiyat giriniz.")
            return
            
        selected_type = self.receipt_type.get()
        if selected_type not in self.calculated_data:
            messagebox.showwarning("Uyarı", "Seçili ödeme türü için hesaplama bulunamadı.")
            return
            
        c_name = self.cust_name.get().upper() or ".............."
        p_name = self.prod_name.get().upper()
        if not p_name:
            messagebox.showwarning("Uyarı", "Lütfen Ürün Adı / Kodu giriniz.")
            return
            
        u_type = self.unit_type.get()
        q = self.get_float(self.qty)
        
        price_info = self.calculated_data[selected_type]
        unit_price = price_info["unit_price"]
        total_price = price_info["total_price"]
        profit = price_info["profit"] if self.current_role == "Yönetici" else 0.0
        
        try:
            tarih = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if is_client_mode():
                api_call("post", "/api/satislar", json={
                    "tarih": tarih,
                    "kullanici": self.current_user,
                    "musteri_adi": c_name,
                    "urun_adi": p_name,
                    "miktar": q,
                    "birim": u_type,
                    "fiyat_birimi": self.base_price_type.get(),
                    "torba_agirligi": self.get_float(self.bag_weight),
                    "alis_fiyati": self.get_float(self.purchase_price) if self.current_role == "Yönetici" else 0.0,
                    "baz_satis_fiyati": self.get_float(self.base_price),
                    "odeme_turu": selected_type,
                    "vade_ay": int(self.get_float(self.vade_months)),
                    "vade_orani": self.get_float(self.vade_rate),
                    "birim_fiyat": unit_price,
                    "toplam_tutar": total_price,
                    "kar": profit,
                    "irsaliye_yolu": ""
                })
            else:
                conn = db_connect()
                cursor = conn.cursor()
                cursor.execute("""
                INSERT INTO satislar (
                    tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi,
                    alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar, irsaliye_yolu
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    tarih, self.current_user, c_name, p_name, q, u_type, self.base_price_type.get(), self.get_float(self.bag_weight),
                    self.get_float(self.purchase_price) if self.current_role == "Yönetici" else 0.0,
                    self.get_float(self.base_price), selected_type, int(self.get_float(self.vade_months)), self.get_float(self.vade_rate),
                    unit_price, total_price, profit, ""
                ))
                conn.commit()
                conn.close()
            
            messagebox.showinfo("Başarılı", "Satış kaydı başarıyla veritabanına kaydedildi.")
            self.load_recent_movements()
            self.update_history_products_combobox()
        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt sırasında hata oluştu:\n{str(e)}")

    # --- Satış Çıktısı Al (PDF) ---
    def export_pdf_only(self):
        if not self.calculated_data:
            messagebox.showwarning("Uyarı", "Hesaplanmış bir veri bulunmuyor. Önce fiyat giriniz.")
            return
            
        selected_type = self.receipt_type.get()
        if selected_type not in self.calculated_data:
            messagebox.showwarning("Uyarı", "Seçili ödeme türü için hesaplama bulunamadı.")
            return
            
        c_name = self.cust_name.get().upper() or ".............."
        p_name = self.prod_name.get().upper()
        u_type = self.unit_type.get()
        q = self.get_float(self.qty)
        
        price_info = self.calculated_data[selected_type]
        unit_price = price_info["unit_price"]
        total_price = price_info["total_price"]
        
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")], initialfile=f"Hausmart_Fis_{c_name.strip('. ')}.pdf")
        if not filepath: return
        
        try:
            self.generate_pdf_receipt(filepath, c_name, p_name, u_type, q, unit_price, total_price)
            messagebox.showinfo("Başarılı", f"Teslim fişi başarıyla oluşturuldu:\n{filepath}")
            os.startfile(filepath)
        except Exception as e:
            messagebox.showerror("Hata", f"PDF oluşturulamadı:\n{str(e)}")

    # --- Satışı Kaydet ve Fiş Oluştur ---
    def save_sale_and_print(self):
        if not self.calculated_data:
            messagebox.showwarning("Uyarı", "Hesaplanmış bir veri bulunmuyor. Önce fiyat giriniz.")
            return
            
        selected_type = self.receipt_type.get()
        if selected_type not in self.calculated_data:
            messagebox.showwarning("Uyarı", "Seçili ödeme türü için hesaplama bulunamadı.")
            return
            
        c_name = self.cust_name.get().upper() or ".............."
        p_name = self.prod_name.get().upper()
        if not p_name:
            messagebox.showwarning("Uyarı", "Lütfen Ürün Adı / Kodu giriniz.")
            return
            
        u_type = self.unit_type.get()
        q = self.get_float(self.qty)
        
        price_info = self.calculated_data[selected_type]
        unit_price = price_info["unit_price"]
        total_price = price_info["total_price"]
        profit = price_info["profit"] if self.current_role == "Yönetici" else 0.0
        
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")], initialfile=f"Hausmart_Fis_{c_name.strip('. ')}.pdf")
        if not filepath: return
        
        try:
            self.generate_pdf_receipt(filepath, c_name, p_name, u_type, q, unit_price, total_price)
            
            tarih = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if is_client_mode():
                api_call("post", "/api/satislar", json={
                    "tarih": tarih,
                    "kullanici": self.current_user,
                    "musteri_adi": c_name,
                    "urun_adi": p_name,
                    "miktar": q,
                    "birim": u_type,
                    "fiyat_birimi": self.base_price_type.get(),
                    "torba_agirligi": self.get_float(self.bag_weight),
                    "alis_fiyati": self.get_float(self.purchase_price) if self.current_role == "Yönetici" else 0.0,
                    "baz_satis_fiyati": self.get_float(self.base_price),
                    "odeme_turu": selected_type,
                    "vade_ay": int(self.get_float(self.vade_months)),
                    "vade_orani": self.get_float(self.vade_rate),
                    "birim_fiyat": unit_price,
                    "toplam_tutar": total_price,
                    "kar": profit,
                    "irsaliye_yolu": ""
                })
            else:
                conn = db_connect()
                cursor = conn.cursor()
                cursor.execute("""
                INSERT INTO satislar (
                    tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi,
                    alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar, irsaliye_yolu
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    tarih, self.current_user, c_name, p_name, q, u_type, self.base_price_type.get(), self.get_float(self.bag_weight),
                    self.get_float(self.purchase_price) if self.current_role == "Yönetici" else 0.0,
                    self.get_float(self.base_price), selected_type, int(self.get_float(self.vade_months)), self.get_float(self.vade_rate),
                    unit_price, total_price, profit, ""
                ))
                conn.commit()
                conn.close()
            
            messagebox.showinfo("Başarılı", f"Satış kaydedildi ve teslim fişi başarıyla oluşturuldu:\n{filepath}")
            
            self.load_recent_movements()
            self.update_history_products_combobox()
            
            os.startfile(filepath)
            
        except Exception as e:
            messagebox.showerror("Hata", f"İşlem gerçekleştirilirken hata oluştu:\n{str(e)}")

    # --- PDF Fiş Oluşturma Çekirdek Kodu ---
    def generate_pdf_receipt(self, filepath, c_name, p_name, u_type, q, unit_price, total_price):
        c = canvas.Canvas(filepath, pagesize=A4)
        width, height = A4 # 595.27 x 841.89
        
        c.setFillColorRGB(248/255, 205/255, 36/255)
        c.rect(0, height - 100, width, 100, fill=1, stroke=0)
        
        c.setFillColorRGB(0, 0, 0)
        
        c.setFont(FONT_NAME_BOLD, 14)
        c.drawString(40, height - 30, "YILDIZ ÖZYAPI GEREÇLERİ")
        c.setFont(FONT_NAME_BOLD, 24)
        c.drawString(40, height - 55, "Hausmart")
        
        c.setFont(FONT_NAME, 9)
        info_lines = [
            "Buğday Pazarı Mahallesi, Şehit Selim Çelikel Sokak",
            "Başakkent Sitesi D-E Blok, No : 16 Merkez / ÇANKIRI",
            "Telefon : +90 (376) 212 03 65 - +90 (549) 730 68 95",
            "www.yildizozyapi.com | www.hausmart.com.tr"
        ]
        y_pos = height - 70
        for line in info_lines:
            c.drawString(40, y_pos, line)
            y_pos -= 12

        c.setFont(FONT_NAME_BOLD, 12)
        c.drawRightString(width - 40, height - 30, "ÇANKIRI        VERESİYE TESLİM FİŞİ")
        
        c.setFont(FONT_NAME, 9)
        now = datetime.datetime.now()
        date_str = now.strftime("%H:%M:%S        %d.%m.%Y")
        c.drawRightString(width - 40, height - 50, date_str)
        
        c.drawRightString(width - 40, height - 65, "YILDIZ ÖZYAPI GEREÇLERİ İNŞ. SAN. TİC. LTD. ŞTİ.")
        c.drawRightString(width - 40, height - 77, "TİCARET SİCİL NO : 2-1866 | MERSİS NO : 0985004080500018")
        c.drawRightString(width - 40, height - 89, "VERGİ DAİRESİ : ÇANKIRI | VERGİ NO : 985 004 0805")

        y_pos = height - 130
        c.setFont(FONT_NAME, 10)
        c.drawString(40, y_pos, "SAYIN")
        c.setFont(FONT_NAME_BOLD, 11)
        c.drawString(40, y_pos - 15, c_name)
        
        c.setFont(FONT_NAME, 10)
        c.drawRightString(width - 40, y_pos, "VERGİ DAİRESİ : ÇANKIRI")
        c.drawRightString(width - 40, y_pos - 15, "VERGİ NO :")

        y_pos -= 45
        c.setFillColorRGB(248/255, 205/255, 36/255)
        c.rect(40, y_pos, width - 80, 20, fill=1, stroke=0)
        
        c.setFillColorRGB(0, 0, 0)
        c.setFont(FONT_NAME_BOLD, 9)
        c.drawString(45, y_pos + 6, "ÜRÜN KODU")
        c.drawString(130, y_pos + 6, "ÜRÜN ADI")
        c.drawString(320, y_pos + 6, "MİKTAR")
        c.drawString(380, y_pos + 6, "İSKONTO")
        c.drawString(440, y_pos + 6, "BİRİM")
        c.drawRightString(width - 45, y_pos + 6, "BİRİM FİYAT")

        c.setFont(FONT_NAME, 9)
        y_pos -= 20
        
        for i in range(15):
            if i % 2 == 0:
                c.setFillColorRGB(0.95, 0.95, 0.95)
                c.rect(40, y_pos, width - 80, 20, fill=1, stroke=0)
            
            c.setFillColorRGB(0, 0, 0)
            if i == 0 and p_name:
                qty_str = f"{q:g}"
                u_price_str = f"{unit_price:,.2f}"
                c.drawString(45, y_pos + 6, "-")
                c.setFont(FONT_NAME_BOLD, 9)
                c.drawString(130, y_pos + 6, p_name)
                c.setFont(FONT_NAME, 9)
                c.drawString(320, y_pos + 6, qty_str)
                c.drawString(380, y_pos + 6, "0,00")
                c.drawString(440, y_pos + 6, u_type)
                c.drawRightString(width - 45, y_pos + 6, u_price_str)
            
            y_pos -= 20

        y_pos -= 15
        total_str = f"{total_price:,.2f}"
        
        c.setFont(FONT_NAME, 10)
        c.drawRightString(width - 120, y_pos, "Ara Toplam (Net) :")
        c.drawRightString(width - 45, y_pos, total_str)
        
        c.drawRightString(width - 120, y_pos - 15, "KDV Toplam :")
        c.drawRightString(width - 45, y_pos - 15, "0,00")

        c.setFillColorRGB(0.85, 0.85, 0.85)
        c.rect(width - 250, y_pos - 35, 210, 18, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)
        c.setFont(FONT_NAME_BOLD, 11)
        c.drawRightString(width - 120, y_pos - 30, "TOPLAM :")
        c.drawRightString(width - 45, y_pos - 30, total_str)

        y_pos -= 80
        c.setFont(FONT_NAME, 8)
        c.drawCentredString(width / 2, y_pos, "YUKARIDA ADI VE MİKTARI BELİRTİLEN ÜRÜNLERİ NOKSANSIZ VE TAM OLARAK TESLİM ALDIM.")
        c.drawCentredString(width / 2, y_pos - 12, "BEDELİ BORCUMDUR 15 GÜN İÇERİSİNDE ÖDEYECEĞİM. AKSİ TAKTİRDE ÇANKIRI MAHKEMELERİ VE İCRA DAİRELERİ YETKİLİDİR.")

        y_pos -= 50
        c.setFont(FONT_NAME_BOLD, 9)
        c.drawString(80, y_pos, "EKSİKSİZ TESLİM EDEN")
        c.setFont(FONT_NAME, 9)
        c.drawString(100, y_pos - 15, "İMZA")

        c.setFont(FONT_NAME_BOLD, 9)
        c.drawRightString(width - 80, y_pos, "EKSİKSİZ TESLİM ALAN")
        c.setFont(FONT_NAME, 9)
        c.drawRightString(width - 110, y_pos - 15, "İMZA")

        c.save()

    def recalculate_and_heal_sales_profits(self):
        try:
            conn = db_connect()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            sales = cursor.execute("SELECT * FROM satislar").fetchall()
            for r in sales:
                qty = r["miktar"] if r["miktar"] is not None else 0.0
                if qty <= 0:
                    continue
                base_price = r["baz_satis_fiyati"] if r["baz_satis_fiyati"] is not None else 0.0
                purchase_price = r["alis_fiyati"] if r["alis_fiyati"] is not None else 0.0
                bag_weight = r["torba_agirligi"] if r["torba_agirligi"] is not None else 50.0
                u_qty_type = r["birim"] if r["birim"] is not None else 'TORBA'
                u_pur_type = r["alis_birimi"] if r["alis_birimi"] is not None else u_qty_type
                u_sel_type = r["fiyat_birimi"] if r["fiyat_birimi"] is not None else u_qty_type

                weights = {
                    'KG': 1.0,
                    'TON': 1000.0,
                    'TORBA': bag_weight,
                    'M2': 1.0
                }

                pur_conversion_factor = 1.0
                if u_qty_type != u_pur_type and u_qty_type != 'M2' and u_pur_type != 'M2':
                    w_q = weights.get(u_qty_type, 1.0)
                    w_pur = weights.get(u_pur_type, 1.0)
                    pur_conversion_factor = w_q / w_pur

                sel_conversion_factor = 1.0
                if u_qty_type != u_sel_type and u_qty_type != 'M2' and u_sel_type != 'M2':
                    w_q = weights.get(u_qty_type, 1.0)
                    w_sel = weights.get(u_sel_type, 1.0)
                    sel_conversion_factor = w_q / w_sel

                converted_purchase_price = purchase_price * pur_conversion_factor
                converted_base_price = base_price * sel_conversion_factor
                ship_cost = r["nakliye_maliyeti"] if (r["nakliye_dahil"] == 1 and r["nakliye_maliyeti"] is not None) else 0.0
                unload_cost = r["indirme_maliyeti"] if (r["indirme_dahil"] == 1 and r["indirme_maliyeti"] is not None) else 0.0

                unit_extra_cost = (ship_cost + unload_cost) / qty
                unit_cost = (converted_purchase_price + unit_extra_cost) if converted_purchase_price > 0 else 0.0
                unit_base_price = converted_base_price + unit_extra_cost

                expected_profit = ((unit_base_price - unit_cost) * qty) if unit_cost > 0 else 0.0
                stored_profit = r["kar"] if r["kar"] is not None else 0.0
                diff = abs(expected_profit - stored_profit)

                if diff > 0.02:
                    cursor.execute("UPDATE satislar SET kar=? WHERE id=?", (expected_profit, r["id"]))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Auto-heal in Python client failed: {e}")

    # --- Hareketler Veri Yükleme ve İrsaliye İşlemleri ---
    def load_recent_movements(self):
        if not hasattr(self, 'mov_tree'): return
        for item in self.mov_tree.get_children():
            self.mov_tree.delete(item)
            
        if not is_client_mode():
            self.recalculate_and_heal_sales_profits()
        if is_client_mode():
            try:
                items = api_call("get", "/api/satislar")
                rows = []
                for item in items:
                    if self.current_role == "Yönetici":
                        rows.append((
                            item["id"], item["tarih"], item["kullanici"], item["musteri_adi"],
                            item["urun_adi"], item["miktar"], item["birim"], item["toplam_tutar"],
                            item["kar"], item["irsaliye_yolu"]
                        ))
                    else:
                        rows.append((
                            item["id"], item["tarih"], item["kullanici"], item["musteri_adi"],
                            item["urun_adi"], item["miktar"], item["birim"], item["toplam_tutar"],
                            item["irsaliye_yolu"]
                        ))
            except Exception as e:
                messagebox.showerror("Bağlantı Hatası", f"Sunucudan veriler alınamadı:\n{e}")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            
            if self.current_role == "Yönetici":
                cursor.execute("SELECT id, tarih, kullanici, musteri_adi, urun_adi, miktar, birim, toplam_tutar, kar, irsaliye_yolu FROM satislar ORDER BY id DESC")
            else:
                cursor.execute("SELECT id, tarih, kullanici, musteri_adi, urun_adi, miktar, birim, toplam_tutar, irsaliye_yolu FROM satislar ORDER BY id DESC")
                
            rows = cursor.fetchall()
            conn.close()
        
        for r in rows:
            irs_path = r[-1]
            irs_status = "Yüklendi" if irs_path else "Yüklenmedi"
            
            row_vals = list(r[:-1])
            row_vals.append(irs_status)
            
            row_vals[5] = f"{row_vals[5]:g}"
            if self.current_role == "Yönetici":
                total_tutar = row_vals[7]
                kar = row_vals[8]
                total_cost = total_tutar - kar
                profit_pct = (kar / total_cost * 100) if total_cost > 0 else 0.0
                pct_sign = "+" if profit_pct >= 0 else ""
                row_vals[8] = f"{kar:,.2f} ({pct_sign}{profit_pct:.2f}%)" if total_cost > 0 else f"{kar:,.2f}"
            row_vals[7] = f"{row_vals[7]:,.2f}"
                
            self.mov_tree.insert("", tk.END, values=row_vals)

    def show_sale_details(self):
        selected = self.mov_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen detayını görüntülemek istediğiniz satışı listeden seçin.")
            return
            
        item_vals = self.mov_tree.item(selected[0], "values")
        satis_id = item_vals[0]
        
        SaleDetailsDialog(self.root, satis_id, self.current_role)

    def upload_waybill(self):
        selected = self.mov_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen irsaliye yüklemek istediğiniz satışı listeden seçin.")
            return
            
        item_vals = self.mov_tree.item(selected[0], "values")
        satis_id = item_vals[0]
        
        file_path = filedialog.askopenfilename(
            title="İrsaliye Dosyası Seçin",
            filetypes=[("Resim & Belge", "*.pdf;*.png;*.jpg;*.jpeg"), ("Tüm Dosyalar", "*.*")]
        )
        if not file_path: return
        
        os.makedirs("irsaliyeler", exist_ok=True)
        _, ext = os.path.splitext(file_path)
        dest_filename = f"irsaliye_{satis_id}{ext}"
        dest_path = os.path.join("irsaliyeler", dest_filename)
        
        try:
            shutil.copy2(file_path, dest_path)
            
            if is_client_mode():
                api_call("put", f"/api/satislar/{satis_id}/irsaliye", json={"yol": dest_path})
            else:
                conn = db_connect()
                cursor = conn.cursor()
                cursor.execute("UPDATE satislar SET irsaliye_yolu = ? WHERE id = ?", (dest_path, satis_id))
                conn.commit()
                conn.close()
            
            messagebox.showinfo("Başarılı", "İrsaliye dosyası başarıyla yüklendi ve bu satışla ilişkilendirildi.")
            self.load_recent_movements()
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya kopyalanırken hata oluştu:\n{str(e)}")

    def view_waybill(self):
        selected = self.mov_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen irsaliyesini görüntülemek istediğiniz satışı listeden seçin.")
            return
            
        item_vals = self.mov_tree.item(selected[0], "values")
        satis_id = item_vals[0]
        
        if is_client_mode():
            try:
                data = api_call("get", f"/api/satislar/{satis_id}")
                path = data.get("irsaliye_yolu", "")
                row = (path,) if path else None
            except Exception as e:
                messagebox.showerror("Hata", f"İrsaliye bilgisi alınamadı: {e}")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("SELECT irsaliye_yolu FROM satislar WHERE id = ?", (satis_id,))
            row = cursor.fetchone()
            conn.close()
        
        if row and row[0]:
            path = row[0]
            if os.path.exists(path):
                os.startfile(path)
            else:
                messagebox.showerror("Hata", f"İrsaliye dosyası bulunamadı:\n{path}")
        else:
            messagebox.showwarning("Bilgi", "Bu satış kaydına ait bir irsaliye yüklenmemiş.")

    def regenerate_pdf_from_history(self):
        selected = self.mov_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen teslim fişini tekrar çıkarmak istediğiniz satışı seçin.")
            return
            
        item_vals = self.mov_tree.item(selected[0], "values")
        satis_id = item_vals[0]
        
        if is_client_mode():
            try:
                data = api_call("get", f"/api/satislar/{satis_id}")
                row = (
                    data.get("musteri_adi"), data.get("urun_adi"), data.get("birim"),
                    data.get("miktar"), data.get("birim_fiyat"), data.get("toplam_tutar")
                )
            except Exception as e:
                messagebox.showerror("Hata", f"Satış bilgisi alınamadı: {e}")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("""
            SELECT musteri_adi, urun_adi, birim, miktar, birim_fiyat, toplam_tutar
            FROM satislar WHERE id = ?
            """, (satis_id,))
            row = cursor.fetchone()
            conn.close()
        
        if not row:
            messagebox.showerror("Hata", "Satış kaydı bulunamadı.")
            return
            
        c_name, p_name, u_type, q, unit_price, total_price = row
        
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")], initialfile=f"Hausmart_Fis_{c_name.strip('. ')}.pdf")
        if not filepath: return
        
        try:
            self.generate_pdf_receipt(filepath, c_name, p_name, u_type, q, unit_price, total_price)
            messagebox.showinfo("Başarılı", f"Teslim fişi başarıyla tekrar oluşturuldu:\n{filepath}")
            os.startfile(filepath)
        except Exception as e:
            messagebox.showerror("Hata", f"PDF oluşturulamadı:\n{str(e)}")

    # --- Seçili Satış Detay PDF ---
    def export_sale_detail_pdf(self):
        selected = self.mov_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen PDF çıktısı almak istediğiniz satışı seçin.")
            return

        item_vals = self.mov_tree.item(selected[0], "values")
        satis_id = item_vals[0]

        if is_client_mode():
            try:
                data = api_call("get", f"/api/satislar/{satis_id}")
                row = (
                    data["tarih"], data["kullanici"], data.get("musteri_adi"), data.get("urun_adi"),
                    data.get("miktar"), data.get("birim"), data.get("fiyat_birimi"), data.get("torba_agirligi"),
                    data.get("alis_fiyati", 0.0), data.get("baz_satis_fiyati", 0.0), data.get("odeme_turu"),
                    data.get("vade_ay", 0), data.get("vade_orani", 0.0), data.get("birim_fiyat", 0.0),
                    data.get("toplam_tutar", 0.0), data.get("kar", 0.0)
                )
            except Exception as e:
                messagebox.showerror("Hata", f"Satış bilgisi alınamadı: {e}")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("""
            SELECT tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi,
                   alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar
            FROM satislar WHERE id = ?
            """, (satis_id,))
            row = cursor.fetchone()
            conn.close()

        if not row:
            messagebox.showerror("Hata", "Satış kaydı bulunamadı.")
            return

        (tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi,
         alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar) = row

        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")],
            initialfile=f"Siparis_Detay_{satis_id}_{(musteri_adi or 'musteri').strip('. ')}.pdf"
        )
        if not filepath: return

        try:
            c = canvas.Canvas(filepath, pagesize=A4)
            width, height = A4

            # Sarı başlık bandı
            c.setFillColorRGB(248/255, 205/255, 36/255)
            c.rect(0, height - 100, width, 100, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
            c.setFont(FONT_NAME_BOLD, 14)
            c.drawString(40, height - 30, "YILDIZ ÖZYAPI GEREÇLERİ")
            c.setFont(FONT_NAME_BOLD, 24)
            c.drawString(40, height - 55, "Hausmart")
            c.setFont(FONT_NAME, 9)
            info_lines = [
                "Buğday Pazarı Mahallesi, Şehit Selim Çelikel Sokak",
                "Başakkent Sitesi D-E Blok, No : 16 Merkez / ÇANKIRI",
                "Telefon : +90 (376) 212 03 65 - +90 (549) 730 68 95",
                "www.yildizozyapi.com | www.hausmart.com.tr"
            ]
            y_pos = height - 70
            for line in info_lines:
                c.drawString(40, y_pos, line)
                y_pos -= 12

            c.setFont(FONT_NAME_BOLD, 12)
            c.drawRightString(width - 40, height - 30, "SİPARİŞ DETAY RAPORU")
            c.setFont(FONT_NAME, 9)
            c.drawRightString(width - 40, height - 50, f"Rapor No: {satis_id}")
            c.drawRightString(width - 40, height - 65, f"Oluşturma: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}") 

            # Detay bloğu
            y = height - 140
            c.setFillColorRGB(248/255, 205/255, 36/255)
            c.rect(40, y, width - 80, 22, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
            c.setFont(FONT_NAME_BOLD, 11)
            c.drawString(50, y + 6, "SİPARİŞ BİLGİLERİ")

            y -= 8
            fields = [
                ("Satış Tarihi", tarih),
                ("Satışı Yapan", kullanici),
                ("Müşteri", musteri_adi or "-"),
                ("Ürün Adı / Kodu", urun_adi or "-"),
                ("Miktar", f"{miktar:g} {birim}"),
                ("Satış Fiyat Birimi", fiyat_birimi),
            ]
            if birim == "TORBA" or fiyat_birimi == "TORBA":
                fields.append(("Torba Ağırlığı", f"{torba_agirligi:g} kg"))
            fields += [
                ("Ödeme Türü", odeme_turu),
                ("Vade", f"{vade_ay} Ay (Aylık %{vade_orani:.2f})" if vade_ay > 0 else "Nakit / Vadesiz"),
                ("Baz Satış Fiyatı", f"{baz_satis_fiyati:,.2f} â‚º"),
                ("Birim Satış Fiyatı", f"{birim_fiyat:,.2f} â‚º"),
                ("Toplam Tutar", f"{toplam_tutar:,.2f} â‚º"),
            ]
            if self.current_role == "Yönetici" and alis_fiyati > 0:
                fields.append(("Alış Fiyatı", f"{alis_fiyati:,.2f} â‚º"))
                fields.append(("Toplam KÃ¢r", f"{kar:,.2f} â‚º"))

            for i, (lbl, val) in enumerate(fields):
                y -= 22
                if i % 2 == 0:
                    c.setFillColorRGB(0.95, 0.95, 0.95)
                    c.rect(40, y - 4, width - 80, 20, fill=1, stroke=0)
                c.setFillColorRGB(0, 0, 0)
                c.setFont(FONT_NAME_BOLD, 10)
                c.drawString(50, y + 3, lbl + ":")
                c.setFont(FONT_NAME, 10)
                c.drawString(250, y + 3, val)

            # Toplam kutusu
            y -= 40
            c.setFillColorRGB(0.85, 0.85, 0.85)
            c.rect(width - 260, y - 5, 220, 22, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
            c.setFont(FONT_NAME_BOLD, 12)
            c.drawRightString(width - 120, y + 3, "GENEL TOPLAM :")
            c.drawRightString(width - 45, y + 3, f"{toplam_tutar:,.2f} â‚º")

            c.save()
            messagebox.showinfo("Başarılı", f"Detay PDF oluşturuldu:\n{filepath}")
            os.startfile(filepath)
        except Exception as e:
            messagebox.showerror("Hata", f"PDF oluşturulamadı:\n{str(e)}")

    # --- Seçili Satış Detay Excel ---
    def export_sale_detail_excel(self):
        selected = self.mov_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen Excel çıktısı almak istediğiniz satışı seçin.")
            return

        item_vals = self.mov_tree.item(selected[0], "values")
        satis_id = item_vals[0]

        if is_client_mode():
            try:
                data = api_call("get", f"/api/satislar/{satis_id}")
                row = (
                    data["tarih"], data["kullanici"], data.get("musteri_adi"), data.get("urun_adi"),
                    data.get("miktar"), data.get("birim"), data.get("fiyat_birimi"), data.get("torba_agirligi"),
                    data.get("alis_fiyati", 0.0), data.get("baz_satis_fiyati", 0.0), data.get("odeme_turu"),
                    data.get("vade_ay", 0), data.get("vade_orani", 0.0), data.get("birim_fiyat", 0.0),
                    data.get("toplam_tutar", 0.0), data.get("kar", 0.0)
                )
            except Exception as e:
                messagebox.showerror("Hata", f"Satış bilgisi alınamadı: {e}")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("""
            SELECT tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi,
                   alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar
            FROM satislar WHERE id = ?
            """, (satis_id,))
            row = cursor.fetchone()
            conn.close()

        if not row:
            messagebox.showerror("Hata", "Satış kaydı bulunamadı.")
            return

        (tarih, kullanici, musteri_adi, urun_adi, miktar, birim, fiyat_birimi, torba_agirligi,
         alis_fiyati, baz_satis_fiyati, odeme_turu, vade_ay, vade_orani, birim_fiyat, toplam_tutar, kar) = row

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Siparis_{satis_id}_{(musteri_adi or 'musteri').strip('. ')}.xlsx"
        )
        if not filepath: return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sipariş Detayı"

            yellow_fill = PatternFill(start_color="F8CD24", end_color="F8CD24", fill_type="solid")
            grey_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            bold = Font(bold=True)
            title_font = Font(bold=True, size=13)
            center = Alignment(horizontal="center")

            ws.merge_cells("A1:C1")
            ws["A1"] = f"HAUSMART - SİPARİŞ DETAYI (ID: {satis_id})"
            ws["A1"].font = title_font
            ws["A1"].fill = yellow_fill
            ws["A1"].alignment = center

            fields = [
                ("Satış Tarihi", tarih),
                ("Satışı Yapan", kullanici),
                ("Müşteri", musteri_adi or "-"),
                ("Ürün Adı / Kodu", urun_adi or "-"),
                ("Miktar", f"{miktar:g} {birim}"),
                ("Satış Fiyat Birimi", fiyat_birimi),
                ("Torba Ağırlığı (kg)", torba_agirligi if (birim == "TORBA" or fiyat_birimi == "TORBA") else "-"),
                ("Ödeme Türü", odeme_turu),
                ("Vade (Ay)", vade_ay),
                ("Aylık Vade Oranı (%)", vade_orani),
                ("Baz Satış Fiyatı (â‚º)", baz_satis_fiyati),
                ("Birim Satış Fiyatı (â‚º)", birim_fiyat),
                ("Toplam Tutar (â‚º)", toplam_tutar),
            ]
            if self.current_role == "Yönetici" and alis_fiyati > 0:
                fields.append(("Alış Fiyatı (â‚º)", alis_fiyati))
                fields.append(("Toplam KÃ¢r (â‚º)", kar))

            for i, (lbl, val) in enumerate(fields, start=2):
                cell_lbl = ws.cell(row=i, column=1, value=lbl)
                cell_lbl.font = bold
                if i % 2 == 0:
                    cell_lbl.fill = grey_fill
                cell_val = ws.cell(row=i, column=2, value=val)
                if i % 2 == 0:
                    cell_val.fill = grey_fill
                if isinstance(val, float):
                    cell_val.number_format = '#,##0.00'

            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 22

            wb.save(filepath)
            messagebox.showinfo("Başarılı", f"Sipariş detayı Excel'e kaydedildi:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel oluşturulurken hata:\n{str(e)}")

    def delete_sale(self):
        selected = self.mov_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Silmek istediğiniz satışı seçin.")
            return
            
        item_vals = self.mov_tree.item(selected[0], "values")
        satis_id = item_vals[0]
        
        confirm = messagebox.askyesno("Kayıt Silme Onayı", f"ID'si {satis_id} olan satış kaydını kalıcı olarak silmek istiyor musunuz?")
        if not confirm: return
        
        try:
            if is_client_mode():
                api_call("delete", f"/api/satislar/{satis_id}")
            else:
                conn = db_connect()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM satislar WHERE id = ?", (satis_id,))
                conn.commit()
                conn.close()
            
            messagebox.showinfo("Başarılı", "Satış kaydı başarıyla silindi.")
            self.load_recent_movements()
            self.update_history_products_combobox()
        except Exception as e:
            messagebox.showerror("Hata", f"Silme işlemi başarısız:\n{str(e)}")

    # --- Ürün Fiyat Geçmişi Yardımcı Fonksiyonları ---
    def update_history_products_combobox(self):
        if not hasattr(self, 'cb_history_products'): return
        if is_client_mode():
            try:
                products = api_call("get", "/api/urunler")
            except Exception as e:
                products = []
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT urun_adi FROM satislar WHERE urun_adi IS NOT NULL AND urun_adi != '' ORDER BY urun_adi ASC")
            rows = cursor.fetchall()
            conn.close()
            products = [r[0] for r in rows]
        
        self.cb_history_products.config(values=products)
        if products:
            self.cb_history_products.set(products[0])
            self.load_price_history()
        else:
            self.cb_history_products.set("")
            if hasattr(self, 'hist_tree'):
                for item in self.hist_tree.get_children():
                    self.hist_tree.delete(item)

    def load_price_history(self):
        if not hasattr(self, 'hist_tree'): return
        for item in self.hist_tree.get_children():
            self.hist_tree.delete(item)
            
        prod = self.cb_history_products.get()
        if not prod: return
        
        if is_client_mode():
            try:
                items = api_call("get", f"/api/fiyat_gecmisi/{prod}")
                rows = []
                for item in items:
                    rows.append((
                        item["tarih"], item["musteri_adi"], item["miktar"], item["birim"],
                        item["odeme_turu"], item["vade_ay"], item["birim_fiyat"], item["toplam_tutar"]
                    ))
            except Exception as e:
                messagebox.showerror("Hata", f"Ürün fiyat geçmişi alınamadı: {e}")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("""
            SELECT tarih, musteri_adi, miktar, birim, odeme_turu, vade_ay, birim_fiyat, toplam_tutar
            FROM satislar WHERE urun_adi = ? ORDER BY tarih DESC
            """, (prod,))
            rows = cursor.fetchall()
            conn.close()
        
        for r in rows:
            row_vals = list(r)
            row_vals[2] = f"{row_vals[2]:g}"
            row_vals[6] = f"{row_vals[6]:,.2f}"
            row_vals[7] = f"{row_vals[7]:,.2f}"
            self.hist_tree.insert("", tk.END, values=row_vals)

    # =========================================================
    # --- Sekme 4: Kullanıcı Yönetimi (Sadece Süper Admin) ---
    # =========================================================
    def create_tab4_widgets(self):
        main_frame = ttk.Frame(self.tab4)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(top_frame, text="Kullanıcı Yönetimi",
                  font=("Helvetica", 13, "bold"), foreground="#1d1d1f").pack(side=tk.LEFT)
        ttk.Button(top_frame, text="Yenile", command=self.load_users).pack(side=tk.RIGHT)

        lf = ttk.LabelFrame(main_frame, text="Kayıtlı Kullanıcılar")
        lf.pack(fill=tk.BOTH, expand=True)

        usr_cols = ("id", "kullanici_adi", "rol", "durum")
        self.usr_tree = ttk.Treeview(lf, columns=usr_cols, show="headings", height=18)
        self.usr_tree.heading("id",           text="ID")
        self.usr_tree.heading("kullanici_adi", text="Kullanıcı Adı")
        self.usr_tree.heading("rol",          text="Rol")
        self.usr_tree.heading("durum",        text="Durum")

        self.usr_tree.column("id",            anchor=tk.CENTER, width=50)
        self.usr_tree.column("kullanici_adi", anchor=tk.W,      width=200)
        self.usr_tree.column("rol",           anchor=tk.CENTER, width=150)
        self.usr_tree.column("durum",         anchor=tk.CENTER, width=100)

        self.usr_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(btn_frame, text="âž•  Kullanıcı Ekle",
                   command=self.add_user_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="âœ   Bilgileri Güncelle",
                   command=self.edit_user_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="ğŸ”‘  Şifre Sıfırla",
                   command=self.reset_password_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="â ¸  Aktif / Pasif",
                   command=self.toggle_user_active).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="ğŸ—‘  Kullanıcıyı Sil",
                   command=self.delete_user_admin).pack(side=tk.RIGHT, padx=5)

        self.load_users()

    def load_users(self):
        if not hasattr(self, 'usr_tree'): return
        for item in self.usr_tree.get_children():
            self.usr_tree.delete(item)

        if is_client_mode():
            try:
                items = api_call("get", "/api/kullanicilar")
                rows = [(item["id"], item["kullanici_adi"], item["rol"], item["aktif"]) for item in items]
            except Exception as e:
                messagebox.showerror("Hata", f"Kullanıcı listesi alınamadı: {e}")
                return
        else:
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("SELECT id, kullanici_adi, rol, aktif FROM kullanicilar ORDER BY id ASC")
            rows = cursor.fetchall()
            conn.close()

        for r in rows:
            uid, adi, rol, aktif = r
            durum = "✅ Aktif" if aktif else "âŒ Pasif"
            tag = "aktif" if aktif else "pasif"
            self.usr_tree.insert("", tk.END, values=(uid, adi, rol, durum), tags=(tag,))

        self.usr_tree.tag_configure("pasif", foreground="#aaaaaa")
        self.usr_tree.tag_configure("aktif", foreground="#1d1d1f")

    def _get_selected_user(self):
        """Seçili kullanıcının (id, adi, rol, aktif) bilgilerini döner."""
        selected = self.usr_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen listeden bir kullanıcı seçin.")
            return None
        vals = self.usr_tree.item(selected[0], "values")
        uid = int(vals[0])
        adi = vals[1]
        rol = vals[2]
        aktif = 1 if "Aktif" in vals[3] else 0
        return uid, adi, rol, aktif

    # ---------- Kullanıcı Ekle ----------
    def add_user_dialog(self):
        win = tk.Toplevel(self.root)
        win.title("Yeni Kullanıcı Ekle")
        win.geometry("360x280")
        win.resizable(False, False)
        win.configure(bg="#f5f5f7")
        win.transient(self.root)
        win.grab_set()
        _center(win, 360, 280)

        frm = ttk.LabelFrame(win, text="Kullanıcı Bilgileri")
        frm.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        ttk.Label(frm, text="Kullanıcı Adı:").grid(row=0, column=0, sticky=tk.W, padx=8, pady=8)
        v_adi   = tk.StringVar()
        ttk.Entry(frm, textvariable=v_adi, width=22).grid(row=0, column=1, padx=8, pady=8)

        ttk.Label(frm, text="Şifre:").grid(row=1, column=0, sticky=tk.W, padx=8, pady=8)
        v_sifre = tk.StringVar()
        ttk.Entry(frm, textvariable=v_sifre, show="*", width=22).grid(row=1, column=1, padx=8, pady=8)

        ttk.Label(frm, text="Rol:").grid(row=2, column=0, sticky=tk.W, padx=8, pady=8)
        v_rol   = tk.StringVar(value="Personel")
        ttk.Combobox(frm, textvariable=v_rol,
                     values=["Süper Admin", "Yönetici", "Personel"],
                     state="readonly", width=19).grid(row=2, column=1, padx=8, pady=8)

        def kaydet():
            adi   = v_adi.get().strip()
            sifre = v_sifre.get().strip()
            rol   = v_rol.get()
            if not adi or not sifre:
                messagebox.showwarning("Eksik Bilgi", "Kullanıcı adı ve şifre boş bırakılamaz.", parent=win)
                return
            try:
                if is_client_mode():
                    try:
                        api_call("post", "/api/kullanicilar", json={"kullanici_adi": adi, "sifre": sifre, "rol": rol})
                    except Exception as e:
                        if "409" in str(e):
                            raise sqlite3.IntegrityError()
                        else:
                            raise e
                else:
                    conn = db_connect()
                    cursor = conn.cursor()
                    cursor.execute(
                        "INSERT INTO kullanicilar (kullanici_adi, sifre, rol, aktif) VALUES (?, ?, ?, 1)",
                        (adi, sifre, rol)
                    )
                    conn.commit()
                    conn.close()
                messagebox.showinfo("Başarılı", f"'{adi}' kullanıcısı eklendi.", parent=win)
                win.destroy()
                self.load_users()
            except sqlite3.IntegrityError:
                messagebox.showerror("Hata", "Bu kullanıcı adı zaten mevcut.", parent=win)
            except Exception as e:
                messagebox.showerror("Hata", f"İşlem başarısız:\n{e}", parent=win)

        bf = ttk.Frame(win)
        bf.pack(pady=5)
        ttk.Button(bf, text="Kaydet", command=kaydet).pack(side=tk.LEFT, padx=8)
        ttk.Button(bf, text="İptal",  command=win.destroy).pack(side=tk.LEFT, padx=8)

    # ---------- Bilgileri Güncelle (ad + rol) ----------
    def edit_user_dialog(self):
        sel = self._get_selected_user()
        if not sel: return
        uid, adi, rol, aktif = sel

        win = tk.Toplevel(self.root)
        win.title(f"Kullanıcı Güncelle â€” {adi}")
        win.geometry("360x220")
        win.resizable(False, False)
        win.configure(bg="#f5f5f7")
        win.transient(self.root)
        win.grab_set()
        _center(win, 360, 220)

        frm = ttk.LabelFrame(win, text="Bilgileri Düzenle")
        frm.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        ttk.Label(frm, text="Yeni Kullanıcı Adı:").grid(row=0, column=0, sticky=tk.W, padx=8, pady=10)
        v_adi = tk.StringVar(value=adi)
        ttk.Entry(frm, textvariable=v_adi, width=22).grid(row=0, column=1, padx=8, pady=10)

        ttk.Label(frm, text="Rol:").grid(row=1, column=0, sticky=tk.W, padx=8, pady=10)
        v_rol = tk.StringVar(value=rol)
        ttk.Combobox(frm, textvariable=v_rol,
                     values=["Süper Admin", "Yönetici", "Personel"],
                     state="readonly", width=19).grid(row=1, column=1, padx=8, pady=10)

        def kaydet():
            yeni_adi = v_adi.get().strip()
            yeni_rol = v_rol.get()
            if not yeni_adi:
                messagebox.showwarning("Eksik Bilgi", "Kullanıcı adı boş bırakılamaz.", parent=win)
                return
            try:
                if is_client_mode():
                    try:
                        api_call("put", f"/api/kullanicilar/{uid}", json={"kullanici_adi": yeni_adi, "rol": yeni_rol})
                    except Exception as e:
                        if "409" in str(e) or "IntegrityError" in str(e) or "500" in str(e):
                            raise sqlite3.IntegrityError()
                        else:
                            raise e
                else:
                    conn = db_connect()
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE kullanicilar SET kullanici_adi = ?, rol = ? WHERE id = ?",
                        (yeni_adi, yeni_rol, uid)
                    )
                    conn.commit()
                    conn.close()
                messagebox.showinfo("Başarılı", "Kullanıcı güncellendi.", parent=win)
                win.destroy()
                self.load_users()
            except sqlite3.IntegrityError:
                messagebox.showerror("Hata", "Bu kullanıcı adı başka bir kullanıcıya ait.", parent=win)
            except Exception as e:
                messagebox.showerror("Hata", f"İşlem başarısız:\n{e}", parent=win)

        bf = ttk.Frame(win)
        bf.pack(pady=5)
        ttk.Button(bf, text="Güncelle", command=kaydet).pack(side=tk.LEFT, padx=8)
        ttk.Button(bf, text="İptal",    command=win.destroy).pack(side=tk.LEFT, padx=8)

    # ---------- Şifre Sıfırla ----------
    def reset_password_dialog(self):
        sel = self._get_selected_user()
        if not sel: return
        uid, adi, rol, aktif = sel

        win = tk.Toplevel(self.root)
        win.title(f"Şifre Sıfırla â€” {adi}")
        win.geometry("340x200")
        win.resizable(False, False)
        win.configure(bg="#f5f5f7")
        win.transient(self.root)
        win.grab_set()
        _center(win, 340, 200)

        frm = ttk.LabelFrame(win, text=f"{adi} â€” Yeni Şifre")
        frm.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        ttk.Label(frm, text="Yeni Şifre:").grid(row=0, column=0, sticky=tk.W, padx=8, pady=10)
        v_s1 = tk.StringVar()
        ttk.Entry(frm, textvariable=v_s1, show="*", width=20).grid(row=0, column=1, padx=8, pady=10)

        ttk.Label(frm, text="Tekrar:").grid(row=1, column=0, sticky=tk.W, padx=8, pady=10)
        v_s2 = tk.StringVar()
        ttk.Entry(frm, textvariable=v_s2, show="*", width=20).grid(row=1, column=1, padx=8, pady=10)

        def kaydet():
            s1 = v_s1.get()
            s2 = v_s2.get()
            if not s1:
                messagebox.showwarning("Eksik", "Şifre boş bırakılamaz.", parent=win)
                return
            if s1 != s2:
                messagebox.showerror("Hata", "Şifreler eşleşmiyor.", parent=win)
                return
            try:
                if is_client_mode():
                    api_call("put", f"/api/kullanicilar/{uid}", json={"sifre": s1})
                else:
                    conn = db_connect()
                    cursor = conn.cursor()
                    cursor.execute("UPDATE kullanicilar SET sifre = ? WHERE id = ?", (s1, uid))
                    conn.commit()
                    conn.close()
                messagebox.showinfo("Başarılı", f"'{adi}' şifresi güncellendi.", parent=win)
                win.destroy()
            except Exception as e:
                messagebox.showerror("Hata", f"Şifre güncellenemedi:\n{e}", parent=win)

        bf = ttk.Frame(win)
        bf.pack(pady=5)
        ttk.Button(bf, text="Kaydet", command=kaydet).pack(side=tk.LEFT, padx=8)
        ttk.Button(bf, text="İptal",  command=win.destroy).pack(side=tk.LEFT, padx=8)

    # ---------- Aktif / Pasif Değiştir ----------
    def toggle_user_active(self):
        sel = self._get_selected_user()
        if not sel: return
        uid, adi, rol, aktif = sel

        if adi == self.current_user:
            messagebox.showwarning("İzin Yok", "Kendi hesabınızı pasif yapamazsınız.")
            return

        yeni = 0 if aktif else 1
        durum_str = "aktif" if yeni else "pasif"
        onay = messagebox.askyesno(
            "Onay",
            f"'{adi}' kullanıcısı {durum_str} yapılsın mı?"
        )
        if not onay: return

        try:
            if is_client_mode():
                api_call("put", f"/api/kullanicilar/{uid}", json={"aktif": yeni})
            else:
                conn = db_connect()
                cursor = conn.cursor()
                cursor.execute("UPDATE kullanicilar SET aktif = ? WHERE id = ?", (yeni, uid))
                conn.commit()
                conn.close()
            self.load_users()
        except Exception as e:
            messagebox.showerror("Hata", f"İşlem başarısız:\n{e}")

    # ---------- Kullanıcı Sil ----------
    def delete_user_admin(self):
        sel = self._get_selected_user()
        if not sel: return
        uid, adi, rol, aktif = sel

        if adi == self.current_user:
            messagebox.showwarning("İzin Yok", "Kendi hesabınızı silemezsiniz.")
            return

        onay = messagebox.askyesno(
            "Silme Onayı",
            f"'{adi}' kullanıcısı kalıcı olarak silinsin mi?\n\nBu işlem geri alınamaz."
        )
        if not onay: return

        try:
            if is_client_mode():
                api_call("delete", f"/api/kullanicilar/{uid}")
            else:
                conn = db_connect()
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM satislar WHERE kullanici = ?", (adi,))
                sales_count = cursor.fetchone()[0]
                if sales_count > 0:
                    messagebox.showerror("Hata", "Bu kullanıcıya ait satış kayıtları bulunmaktadır. Silme işlemine izin verilmez. Kullanıcıyı pasif yapabilirsiniz.")
                    conn.close()
                    return
                cursor.execute("DELETE FROM kullanicilar WHERE id = ?", (uid,))
                conn.commit()
                conn.close()
            messagebox.showinfo("Başarılı", f"'{adi}' kullanıcısı silindi.")
            self.load_users()
        except Exception as e:
            messagebox.showerror("Hata", f"Silme işlemi başarısız:\n{e}")


    # =========================================================
    # --- Bağlantı Ayarları Dialog ---
    # =========================================================
    def show_connection_dialog(self):
        cfg = load_config()
        win = tk.Toplevel(self.root)
        win.title("Baglanti Ayarlari")
        win.geometry("500x400")
        win.resizable(False, False)
        win.configure(bg="#f5f5f7")
        win.transient(self.root)
        win.grab_set()
        _center(win, 500, 400)

        frm = ttk.LabelFrame(win, text="Baglanti Modu Secin")
        frm.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        v_mod = tk.StringVar(value=cfg.get("mod", "yerel"))

        # Yerel
        ttk.Radiobutton(frm, text="Yerel Mod  (sadece bu bilgisayar)",
                        variable=v_mod, value="yerel").pack(anchor=tk.W, padx=10, pady=(8,2))

        # Ag paylasim
        rb_pay = ttk.Radiobutton(frm, text="Ag Paylasim Modu  (ag uzerindeki SQLite dosyasina baglan)",
                                  variable=v_mod, value="paylasim")
        rb_pay.pack(anchor=tk.W, padx=10, pady=2)
        pay_row = ttk.Frame(frm)
        pay_row.pack(fill=tk.X, padx=30, pady=2)
        ttk.Label(pay_row, text="Ag DB Yolu:").pack(side=tk.LEFT)
        v_db = tk.StringVar(value=cfg.get("db_yolu", ""))
        ent_db = ttk.Entry(pay_row, textvariable=v_db, width=32)
        ent_db.pack(side=tk.LEFT, padx=6)
        ttk.Label(frm, text="Ornek: \\\\SUNUCU-PC\\Paylasim\\satis_takip.db",
                  font=("Helvetica", 8), foreground="#86868b").pack(padx=30, pady=(0,6))

        # HTTP istemci
        rb_http = ttk.Radiobutton(frm, text="HTTP Istemci Modu  (satis_server.py calistiran PC'ye baglan)",
                                   variable=v_mod, value="istemci")
        rb_http.pack(anchor=tk.W, padx=10, pady=2)
        http_row = ttk.Frame(frm)
        http_row.pack(fill=tk.X, padx=30, pady=2)
        ttk.Label(http_row, text="Sunucu URL:").pack(side=tk.LEFT)
        v_url = tk.StringVar(value=cfg.get("sunucu_url", "http://192.168.1.100:8765"))
        ttk.Entry(http_row, textvariable=v_url, width=28).pack(side=tk.LEFT, padx=6)
        ttk.Label(frm, text="Ornek: http://192.168.1.100:8765",
                  font=("Helvetica", 8), foreground="#86868b").pack(padx=30, pady=(0,6))

        # Talimatlar
        lf_info = ttk.LabelFrame(win, text="Kurulum Rehberi")
        lf_info.pack(fill=tk.X, padx=20, pady=(0,8))
        ttk.Label(lf_info,
            text=("Ag Paylasim: Sunucu PC klasorunu agin ve UNC yolu girin\n"
                  "HTTP Istemci: Sunucu PC'de  python satis_server.py  calistirin\n"
                  "Test butonuyla baglantıyı dogrulayın, uygulamayı yeniden baslatın"),
            font=("Helvetica", 8), foreground="#555", justify=tk.LEFT).pack(padx=8, pady=4)

        def test_baglanti():
            mod = v_mod.get()
            if mod == "yerel":
                messagebox.showinfo("Bilgi", "Yerel mod secildi, test gerekmez.", parent=win)
                return
            if mod == "paylasim":
                import os as _os
                db_path = v_db.get().strip()
                if _os.path.exists(_os.path.dirname(db_path)):
                    messagebox.showinfo("Baglanti Basarili",
                        f"Ag yoluna erisilebiliyor:\n{_os.path.dirname(db_path)}", parent=win)
                else:
                    messagebox.showerror("Hata",
                        f"Ag yoluna erisilemedı:\n{db_path}\n\nKlasorun paylasildigini ve ag baglantisi oldugunuzu kontrol edin.",
                        parent=win)
                return
            if not REQUESTS_OK:
                messagebox.showerror("Hata",
                    "requests kutuphanesi yuklu degil.\nKomut: pip install requests", parent=win)
                return
            try:
                url = v_url.get().rstrip("/") + "/api/durum"
                r = _requests.get(url, timeout=5)
                data = r.json()
                messagebox.showinfo("Baglanti Basarili",
                    f"Sunucu aktif!\nZaman: {data.get('zaman','-')}\nDB: {data.get('db','-')}",
                    parent=win)
            except Exception as e:
                messagebox.showerror("Baglanti Hatasi",
                    f"Sunucuya baglanamadı:\n{e}", parent=win)

        def kaydet():
            mod = v_mod.get()
            import os as _os
            db_yolu = v_db.get().strip() if mod == "paylasim" else cfg.get("db_yolu", "")
            # Yerel modda varsayilan db
            if mod == "yerel":
                db_yolu = _os.path.join(_BASE_DIR, "satis_takip.db")
            yeni = {
                "mod": mod,
                "db_yolu": db_yolu,
                "sunucu_url": v_url.get().strip()
            }
            save_config(yeni)
            messagebox.showinfo("Kaydedildi",
                "Ayarlar kaydedildi.\nUygulamayı yeniden baslatmaniz onerılir.",
                parent=win)
            if hasattr(self, "lbl_conn_mode"):
                labels = {
                    "yerel":    "Yerel Mod",
                    "paylasim": "Ag Paylasim: " + db_yolu,
                    "istemci":  "HTTP Istemci: " + yeni["sunucu_url"]
                }
                self.lbl_conn_mode.config(text=labels.get(mod, mod))
            win.destroy()

        bf = ttk.Frame(win)
        bf.pack(pady=6)
        ttk.Button(bf, text="Baglantıyı Test Et", command=test_baglanti).pack(side=tk.LEFT, padx=8)
        ttk.Button(bf, text="Kaydet", command=kaydet).pack(side=tk.LEFT, padx=8)
        ttk.Button(bf, text="Iptal",  command=win.destroy).pack(side=tk.LEFT, padx=8)


    def create_tab5_widgets(self):
        main_frame = ttk.Frame(self.tab5)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        ttk.Label(main_frame, text="Sunucu & Ağ Yönetimi",
                  font=("Helvetica", 13, "bold"), foreground="#1d1d1f").pack(anchor=tk.W, pady=(0,12))

        cfg = load_config()
        mod     = cfg.get("mod","yerel")
        db_path = cfg.get("db_yolu","satis_takip.db")
        srv_url = cfg.get("sunucu_url","")

        lf_durum = ttk.LabelFrame(main_frame, text="Mevcut Bağlantı Durumu")
        lf_durum.pack(fill=tk.X, pady=(0,12))

        mod_str = "ð İstemci Modu" if mod == "istemci" else "ð¥  Yerel Mod"
        ttk.Label(lf_durum, text=f"Mod: {mod_str}",
                  font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky=tk.W, padx=8, pady=5)
        ttk.Label(lf_durum, text=f"DB Dosyası: {db_path}",
                  font=("Helvetica", 9)).grid(row=1, column=0, sticky=tk.W, padx=8, pady=3)
        if mod == "istemci":
            ttk.Label(lf_durum, text=f"Sunucu URL: {srv_url}",
                      font=("Helvetica", 9), foreground="#0071e3").grid(
                row=2, column=0, sticky=tk.W, padx=8, pady=3)

        ttk.Button(lf_durum, text="â  Bağlantı Ayarlarını Değiştir",
                   command=self.show_connection_dialog).grid(row=0, column=1, padx=12, pady=5)

        lf_ip = ttk.LabelFrame(main_frame,
                               text="Bu Bilgisayarin Ağ Adresleri (sunucu olarak kullanılacaksa)")
        lf_ip.pack(fill=tk.X, pady=(0,12))

        ips = self._get_local_ips()
        for ip in ips:
            url_str = f"http://{ip}:8765"
            row = ttk.Frame(lf_ip)
            row.pack(fill=tk.X, padx=8, pady=3)
            ttk.Label(row, text=f"ð  {url_str}",
                      font=("Courier", 10, "bold"), foreground="#0071e3").pack(side=tk.LEFT)
            ttk.Button(row, text="Kopyala", width=8,
                       command=lambda u=url_str: self._copy_to_clipboard(u)).pack(side=tk.LEFT, padx=8)

        ttk.Label(lf_ip,
                  text="İstemci bilgisayarlar yukarıdaki adresi 'Bağlantı Ayarları' â† İstemci Modu bölümüne girecek.",
                  font=("Helvetica", 8), foreground="#86868b").pack(padx=8, pady=(0,5))

        lf_kurulum = ttk.LabelFrame(main_frame, text="Sunucu Kurulum Talimatları")
        lf_kurulum.pack(fill=tk.X, pady=(0,12))

        steps = [
            "1  Sunucu olarak kullanılacak bilgisayarda:  python satis_server.py  çalıştırın",
            "2  Açılan panelde Sunucuyu Başlat butonuna tıklayın",
            "3  Yukarıdaki IP adreslerinden birini kopyalayın (örn: http://192.168.1.100:8765)",
            "4  Diğer bilgisayarlarda: Bağlantı â† İstemci Modu â† URL yapıştırın â† Kaydet",
            "5  Bağlantıyı Test Et ile bağlantıyı doğrulayın, ardından uygulamayı yeniden başlatın",
        ]
        for step in steps:
            ttk.Label(lf_kurulum, text=step, font=("Helvetica", 9),
                      justify=tk.LEFT, wraplength=700).pack(anchor=tk.W, padx=8, pady=2)

        test_frame = ttk.Frame(main_frame)
        test_frame.pack(fill=tk.X, pady=5)
        ttk.Button(test_frame, text="Bağlantıyı Test Et",
                   command=self.show_connection_dialog).pack(side=tk.LEFT)
        ttk.Button(test_frame, text="DB Dosyasını Göster",
                   command=lambda: os.startfile(
                       os.path.dirname(load_config().get("db_yolu",
                           os.path.join(_BASE_DIR, "satis_takip.db")))
                   )).pack(side=tk.LEFT, padx=10)
        ttk.Button(test_frame, text="Sunucu Panelini Aç",
                   command=self._open_server_panel).pack(side=tk.LEFT)

    def _get_local_ips(self):
        ips = []
        try:
            hostname = socket.gethostname()
            for info in socket.getaddrinfo(hostname, None):
                ip = info[4][0]
                if ip not in ips and not ip.startswith("127.") and ":" not in ip:
                    ips.append(ip)
        except Exception:
            pass
        return ips or ["127.0.0.1"]

    def _copy_to_clipboard(self, text):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("Kopyalandı", f"Panoya kopyalandı:\n{text}")

    def _open_server_panel(self):
        server_script = os.path.join(_BASE_DIR, "satis_server.py")
        if os.path.exists(server_script):
            import subprocess
            subprocess.Popen(["python", server_script], cwd=_BASE_DIR)
        else:
            messagebox.showerror("Hata", f"satis_server.py bulunamadı:\n{server_script}")

    def load_payment_rates(self):
        try:
            if is_client_mode():
                rates = api_call("get", "/api/ayarlar")
            else:
                conn = db_connect()
                cursor = conn.cursor()
                cursor.execute("SELECT anahtar, deger FROM ayarlar")
                rows = cursor.fetchall()
                conn.close()
                rates = {row[0]: row[1] for row in rows}
            
            if rates:
                if isinstance(rates, list):
                    rates_dict = {item["anahtar"]: item["deger"] for item in rates}
                else:
                    rates_dict = rates
                
                focused = self.root.focus_get()
                mapping = [
                    ("rate_cash", self.rate_cash, self.rate_widgets.get("cash")),
                    ("rate_cc", self.rate_cc, self.rate_widgets.get("cc")),
                    ("rate_check", self.rate_check, self.rate_widgets.get("check")),
                    ("rate_note", self.rate_note, self.rate_widgets.get("note")),
                    ("rate_doc", self.rate_doc, self.rate_widgets.get("doc")),
                    ("rate_dbs", self.rate_dbs, self.rate_widgets.get("dbs"))
                ]
                
                for key, var, widget in mapping:
                    val = rates_dict.get(key)
                    if val is not None:
                        if focused != widget:
                            var.set(str(val))
        except Exception as e:
            print("Rates could not be loaded:", e)

    def save_rates(self):
        rates = {
            "rate_cash": self.rate_cash.get().strip() or "0",
            "rate_cc": self.rate_cc.get().strip() or "0",
            "rate_check": self.rate_check.get().strip() or "0",
            "rate_note": self.rate_note.get().strip() or "0",
            "rate_doc": self.rate_doc.get().strip() or "0",
            "rate_dbs": self.rate_dbs.get().strip() or "0"
        }
        try:
            if is_client_mode():
                api_call("post", "/api/ayarlar", json=rates)
            else:
                conn = db_connect()
                cursor = conn.cursor()
                for k, v in rates.items():
                    cursor.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES (?, ?)", (k, v))
                conn.commit()
                conn.close()
        except Exception as e:
            print("Rates could not be saved:", e)

    def periodic_rate_sync(self):
        if hasattr(self, 'root') and self.current_user:
            self.load_payment_rates()
            self.root.after(5000, self.periodic_rate_sync)

    def check_python_update(self):
        import urllib.request
        cfg = load_config()
        repo = cfg.get("github_repo", "hausmart-dev/satis-takip")
        url = f"https://raw.githubusercontent.com/{repo}/main/package.json"
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5) as response:
                data = json.loads(response.read().decode('utf-8'))
                remote_version = data.get("version", "1.0.0")
                local_version = "1.0.0"
                if remote_version != local_version:
                    self.root.after(0, lambda: self.show_python_update_ui(remote_version))
        except Exception as e:
            print("Python check update failed:", e)

    def show_python_update_ui(self, remote_version):
        self.lbl_update_status.config(text=f"Yeni Sürüm Mevcut: v{remote_version} (Mevcut: v1.0.0)")
        self.update_frame.pack(fill=tk.X, pady=10)
        self.btn_update.pack()

    def install_python_update(self):
        self.btn_update.config(state="disabled", text="Güncelleniyor...")
        threading.Thread(target=self.run_python_update, daemon=True).start()

    def run_python_update(self):
        import urllib.request
        import zipfile
        import shutil
        cfg = load_config()
        repo = cfg.get("github_repo", "hausmart-dev/satis-takip")
        zip_url = f"https://github.com/{repo}/archive/refs/heads/main.zip"
        try:
            zip_path = "update.zip"
            req = urllib.request.Request(zip_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=30) as response, open(zip_path, 'wb') as out_file:
                shutil.copyfileobj(response, out_file)
            
            temp_extract = "update_temp"
            if os.path.exists(temp_extract):
                shutil.rmtree(temp_extract, ignore_errors=True)
            os.makedirs(temp_extract, exist_ok=True)
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_extract)
                
            folders = os.listdir(temp_extract)
            if not folders:
                raise Exception("Archive empty")
            inner_folder = os.path.join(temp_extract, folders[0])
            
            for item in os.listdir(inner_folder):
                if item in ['satis_takip.db', 'config.json', 'node_modules', 'node_temp', 'update.zip', 'update_temp']:
                    continue
                src = os.path.join(inner_folder, item)
                dest = os.path.join(".", item)
                
                if os.path.isdir(src):
                    if os.path.exists(dest):
                        shutil.rmtree(dest, ignore_errors=True)
                    shutil.copytree(src, dest)
                else:
                    shutil.copy2(src, dest)
            
            os.remove(zip_path)
            shutil.rmtree(temp_extract, ignore_errors=True)
            
            self.root.after(0, lambda: messagebox.showinfo("Başarılı", "Güncelleme başarıyla tamamlandı. Lütfen programı yeniden başlatın."))
            self.root.after(0, self.root.destroy)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Hata", f"Güncelleme başarısız:\n{e}"))
            self.root.after(0, lambda: self.btn_update.config(state="normal", text="Son Sürüme Güncelle"))


# --- Pencere ortalama yardımcı fonksiyonu ---
def _center(win, w, h):
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{int(sw/2 - w/2)}+{int(sh/2 - h/2)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = HausmartApp(root)
    root.mainloop()
